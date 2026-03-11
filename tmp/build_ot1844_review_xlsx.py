from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

rows = [
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2650','Corte y Fabricación de Vigas y Perfil Z','TK_Complete',162.0,162.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2660','Corte y Fabricación de Consoldas Exteriores','TK_Complete',36.0,36.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2670','Corte y Fabricación de Consoldas Centrales','TK_Complete',36.0,36.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2680','Corte y Fabricación Distribuidor 8"','TK_Complete',36.0,36.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2690','Armado y Remates de Soldadura Bastidor','TK_Complete',54.0,54.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2710','Inspección de soldadura (VT, LP si aplica)','TK_Complete',9.0,9.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2720','Granallado / Preparación de Superficie','TK_Complete',18.0,18.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bastidor','A2730','Aplicación Capa Imprimante, Enlace y Terminación','TK_Active',54.0,8.1,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Perimetral (12 Un)','A2750','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Perimetral (12 Un)','A2760','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Perimetral (12 Un)','A2770','Armado, Soldadura y Rectificado de Bandejas','TK_Active',108.0,75.6,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Central (12 Un)','A2890','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Central (12 Un)','A2900','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Central (12 Un)','A2910','Armado, Soldadura y Rectificado de Bandejas','TK_Complete',108.0,108.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Bandeja Central (12 Un)','A2920','Terminaciones superficiales (Limpieza, Decapado, Pasivado)','TK_Active',54.0,27.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Barandas (16 Un)','A2800','Fabricación Baranda radial removible','TK_Active',189.0,141.8,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Parrilla de Piso (36 Un)','A2830','Trazado y Corte de FRP','TK_Active',54.0,32.4,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Manifold Distribuidor','A2850','Fabricación del cuerpo central (spool/cabezal)','TK_Complete',36.0,36.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Manifold Distribuidor','A2860','Fabricación Brazos radiales (spools repetitivos)','TK_Complete',18.0,18.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Manifold Distribuidor','A2870','Fabricación Conjunto válvula/boquilla (DET. 1-5)','TK_Complete',18.0,18.0,''],
    ['Entrega 1','TAG 2225-ZM-(02 Unidades)-1','Pre-armado en taller / Despacho','A2940','Bastidor / estructura / Bandejas / Barandas / Parrillas FRP','TK_Active',108.0,21.6,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bastidor','A2980','Corte y Fabricación de Vigas y Perfil Z','TK_Complete',162.0,162.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bastidor','A2990','Corte y Fabricación de Consoldas Exteriores','TK_Complete',36.0,36.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bastidor','A3000','Corte y Fabricación de Consoldas Centrales','TK_Complete',36.0,36.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bastidor','A3010','Corte y Fabricación Distribuidor 8"','TK_Complete',36.0,36.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bandeja Perimetral (12 Un)','A3070','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bandeja Perimetral (12 Un)','A3080','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bandeja Central (12 Un)','A3210','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Bandeja Central (12 Un)','A3220','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Barandas (16 Un)','A3120','Fabricación Baranda radial removible','TK_Active',189.0,141.8,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Parrilla de Piso (36 Un)','A3150','Trazado y Corte de FRP','TK_Active',54.0,32.4,''],
    ['Entrega 2','TAG 2225-ZM-(02 Unidades)-2','Manifold Distribuidor','A3170','Fabricación del cuerpo central (spool/cabezal)','TK_Active',36.0,28.8,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bastidor','A3280','Corte y Fabricación de Vigas y Perfil Z','TK_Active',162.0,81.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bastidor','A3290','Corte y Fabricación de Consoldas Exteriores','TK_Active',36.0,18.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bastidor','A3300','Corte y Fabricación de Consoldas Centrales','TK_Active',36.0,18.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bastidor','A3310','Corte y Fabricación Distribuidor 8"','TK_Complete',36.0,36.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bandeja Perimetral (12 Un)','A3370','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bandeja Perimetral (12 Un)','A3380','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bandeja Central (12 Un)','A3510','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Bandeja Central (12 Un)','A3520','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Barandas (16 Un)','A3420','Fabricación Baranda radial removible','TK_Active',189.0,94.5,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Parrilla de Piso (36 Un)','A3450','Trazado y Corte de FRP','TK_Active',54.0,37.8,''],
    ['Entrega 3','TAG 2225-ZM-(02 Unidades)-3','Manifold Distribuidor','A3470','Fabricación del cuerpo central (spool/cabezal)','TK_Active',36.0,28.8,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bastidor','A3580','Corte y Fabricación de Vigas y Perfil Z','TK_Active',162.0,81.0,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bastidor','A3590','Corte y Fabricación de Consoldas Exteriores','TK_Active',36.0,10.8,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bastidor','A3600','Corte y Fabricación de Consoldas Centrales','TK_Active',36.0,10.8,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bastidor','A3610','Corte y Fabricación Distribuidor 8"','TK_Complete',36.0,36.0,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bandeja Perimetral (12 Un)','A3670','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bandeja Perimetral (12 Un)','A3680','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bandeja Central (12 Un)','A3810','Corte, perforado y conformado (fabricación del cuerpo) SS316L','TK_Complete',72.0,72.0,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Bandeja Central (12 Un)','A3820','Cilindrado (fabricación del cuerpo) SS316L','TK_Complete',36.0,36.0,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Barandas (16 Un)','A3720','Fabricación Baranda radial removible','TK_Active',189.0,94.5,''],
    ['Entrega 4','TAG 2225-ZM-(02 Unidades)-4','Manifold Distribuidor','A3770','Fabricación del cuerpo central (spool/cabezal)','TK_Active',36.0,28.8,''],
    ['General','Sin despacho específico','General','A1640','Planos de Fabricación','TK_Complete',360.0,360.0,''],
    ['General','Sin despacho específico','General','A1660','Suminstro de Materiales','TK_Complete',180.0,180.0,''],
]

out = r"C:\Users\josej\.openclaw\workspace\projects\OT-1844\docs\schedule\OT1844_W012_revision_avance_corte_2026-03-15.xlsx"
wb = Workbook()
ws = wb.active
ws.title = 'Revision_Avance_W012'
headers = ['Entrega','Despacho/Tag','WBS','Actividad ID','Actividad','Estado P6','BAC HH','EV HH','Avance %','Marcar avance','Observaciones']
ws.append(headers)
for entrega, tag, wbs, act_id, actividad, estado, bac, ev, obs in rows:
    avance = ev / bac if bac else 0
    ws.append([entrega, tag, wbs, act_id, actividad, estado, bac, ev, avance, '', obs])

for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(horizontal='center', vertical='center')

widths = {
    'A': 12, 'B': 30, 'C': 34, 'D': 14, 'E': 58, 'F': 14,
    'G': 10, 'H': 10, 'I': 10, 'J': 16, 'K': 28
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for row in ws.iter_rows(min_row=2, min_col=7, max_col=9):
    row[0].number_format = '0.0'
    row[1].number_format = '0.0'
    row[2].number_format = '0.0%'

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:K{ws.max_row}"

tab = Table(displayName='TablaRevisionAvanceW012', ref=f"A1:K{ws.max_row}")
style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

for r in range(2, ws.max_row + 1):
    estado = ws[f'F{r}'].value
    if estado == 'TK_Active':
        for c in range(1, 12):
            ws.cell(r, c).fill = PatternFill('solid', fgColor='FFF2CC')
    elif estado == 'TK_Complete':
        for c in range(1, 12):
            ws.cell(r, c).fill = PatternFill('solid', fgColor='E2F0D9')

wb.save(out)
print(out)