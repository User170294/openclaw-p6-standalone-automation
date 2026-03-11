import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment

DB = r"C:\Users\josej\OneDrive\Documentos\PPMDBSQLite_20221109_BBDD_JJC_Rev B.db"
OUT = r"C:\Users\josej\.openclaw\workspace\projects\OT-1844\docs\schedule\OT1844_W012_revision_avance_corte_2026-03-15_v2.xlsx"
PROJ = 26432
CUTOFF = datetime.strptime('2026-03-15 23:59:00', '%Y-%m-%d %H:%M:%S')

con = sqlite3.connect(DB, timeout=10)
cur = con.cursor()

# WBS map
wbs_rows = cur.execute("select wbs_id, parent_wbs_id, wbs_short_name, wbs_name from PROJWBS where proj_id=?", (PROJ,)).fetchall()
wbs = {r[0]: {'parent': r[1], 'short': r[2], 'name': r[3]} for r in wbs_rows}

def climb(wbs_id):
    out=[]; seen=set()
    while wbs_id and wbs_id in wbs and wbs_id not in seen:
        seen.add(wbs_id)
        n=wbs[wbs_id]
        out.append((n['short'] or '', n['name'] or ''))
        wbs_id=n['parent']
    return list(reversed(out))

def entrega_tag(parts):
    tag=''
    for short,name in parts:
        txt=f"{short} {name}"
        if 'TAG 2225' in txt:
            tag=txt.strip()
            break
    entrega='General'
    if tag:
        if '-1' in tag: entrega='Entrega 1'
        elif '-2' in tag: entrega='Entrega 2'
        elif '-3' in tag: entrega='Entrega 3'
        elif '-4' in tag: entrega='Entrega 4'
        elif '-5' in tag: entrega='Entrega 5'
    return entrega, tag if tag else 'Sin despacho específico'

q = """
select t.task_code, t.task_name, t.status_code, t.target_start_date, t.target_end_date,
       t.act_start_date, t.act_end_date, t.complete_pct_type, t.clndr_id, c.clndr_name, t.wbs_id,
       coalesce(sum(case when tr.rsrc_type='RT_Labor' then tr.target_qty else 0 end),0) as bac_hh,
       coalesce(sum(case when tr.rsrc_type='RT_Labor' then tr.act_reg_qty else 0 end),0) as ev_hh,
       coalesce(sum(case when tr.rsrc_type='RT_Labor' then tr.remain_qty else 0 end),0) as etc_hh
from TASK t
left join TASKRSRC tr on tr.task_id=t.task_id and tr.proj_id=t.proj_id
left join CALENDAR c on c.clndr_id=t.clndr_id
where t.proj_id=?
group by t.task_id, t.task_code, t.task_name, t.status_code, t.target_start_date, t.target_end_date,
         t.act_start_date, t.act_end_date, t.complete_pct_type, t.clndr_id, c.clndr_name, t.wbs_id
order by t.target_start_date, t.task_code
"""
rows = cur.execute(q, (PROJ,)).fetchall()

selected=[]
for r in rows:
    task_code, task_name, status, ts, te, acs, ace, pct_type, clndr_id, clndr_name, wbs_id, bac_hh, ev_hh, etc_hh = r
    if not ts:
        continue
    ts_dt = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
    te_dt = datetime.strptime(te, '%Y-%m-%d %H:%M:%S') if te else ts_dt
    if ts_dt <= CUTOFF and bac_hh > 0:
        parts = climb(wbs_id)
        entrega, tag = entrega_tag(parts)
        wbs_name = parts[-1][1] if parts else ''
        selected.append([
            entrega, tag, wbs_name, task_code, task_name, status,
            ts, te or '', acs or '', ace or '', pct_type,
            clndr_id, clndr_name or '', float(bac_hh), float(ev_hh), float(etc_hh),
            (float(ev_hh) / float(bac_hh) if bac_hh else 0.0), '', ''
        ])

wb = Workbook()
ws = wb.active
ws.title = 'Revision_Avance_W012'
headers = [
    'Entrega','Despacho/Tag','WBS','Actividad ID','Actividad','Estado P6',
    'Inicio Plan','Fin Plan','Inicio Real','Fin Real','% Complete Type',
    'Calendar ID','Calendario','BAC HH','EV HH','ETC HH','Avance acumulado corte %',
    '% avance a cargar','Fecha término si 100%'
]
ws.append(headers)
for row in selected:
    ws.append(row)

for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(horizontal='center', vertical='center')

widths = {'A':12,'B':30,'C':32,'D':14,'E':58,'F':14,'G':18,'H':18,'I':18,'J':18,'K':16,'L':12,'M':20,'N':10,'O':10,'P':10,'Q':12,'R':16,'S':20}
for col,width in widths.items():
    ws.column_dimensions[col].width = width

for r in range(2, ws.max_row+1):
    for c in ('N','O','P'):
        ws[f'{c}{r}'].number_format = '0.0'
    ws[f'Q{r}'].number_format = '0.0%'
    estado = ws[f'F{r}'].value
    if estado == 'TK_Active':
        fill = PatternFill('solid', fgColor='FFF2CC')
    elif estado == 'TK_Complete':
        fill = PatternFill('solid', fgColor='E2F0D9')
    else:
        fill = PatternFill('solid', fgColor='FCE4D6')
    for c in range(1, 20):
        ws.cell(r,c).fill = fill

ws.freeze_panes = 'A2'
ref = f"A1:S{ws.max_row}"
tab = Table(displayName='TablaRevisionAvanceW012v2', ref=ref)
style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)
wb.save(OUT)
print(f'ROWS={len(selected)}')
print(OUT)