from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment

src = r"C:\Users\josej\.openclaw\workspace\projects\OT-1844\docs\schedule\OT1844_W012_revision_avance_corte_2026-03-15_v2.xlsx"
out = r"C:\Users\josej\.openclaw\workspace\projects\OT-1844\docs\schedule\OT1844_W012_revision_avance_corte_2026-03-15_v3.xlsx"

wb = load_workbook(src)
ws = wb['Revision_Avance_W012']

# Insert new column before 'Fecha término si 100%'
ws.insert_cols(19)
ws['S1'] = 'Fecha inicio real a cargar'
ws['T1'] = 'Fecha término si 100%'

# Shifted former last header if needed
if ws['U1'].value is None:
    pass

for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(horizontal='center', vertical='center')

widths = {'A':12,'B':30,'C':32,'D':14,'E':58,'F':14,'G':18,'H':18,'I':18,'J':18,'K':16,'L':12,'M':20,'N':10,'O':10,'P':10,'Q':12,'R':16,'S':20,'T':20}
for col,width in widths.items():
    ws.column_dimensions[col].width = width

# Remove existing tables and recreate
for name in list(ws.tables.keys()):
    del ws.tables[name]
ref = f"A1:T{ws.max_row}"
tab = Table(displayName='TablaRevisionAvanceW012v3', ref=ref)
style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)
ws.freeze_panes = 'A2'

wb.save(out)
print(out)