from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment

src = r"C:\Users\josej\.openclaw\workspace\projects\OT-1844\docs\schedule\OT1844_W012_revision_avance_corte_2026-03-15_v2.xlsx"
out = r"C:\Users\josej\.openclaw\workspace\projects\OT-1844\docs\schedule\OT1844_W012_revision_avance_corte_2026-03-15_v3.xlsx"

wb = load_workbook(src)
ws = wb['Revision_Avance_W012']

# rebuild sheet with correct structure and formatting
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]

new_headers = headers[:18] + ['Fecha inicio real a cargar'] + headers[18:]
new_rows = [row[:18] + [''] + row[18:] for row in rows]

ws.delete_rows(1, ws.max_row)
ws.append(new_headers)
for row in new_rows:
    ws.append(row)

header_fill = PatternFill('solid', fgColor='1F4E78')
complete_fill = PatternFill('solid', fgColor='E2F0D9')
active_fill = PatternFill('solid', fgColor='FFF2CC')
other_fill = PatternFill('solid', fgColor='FCE4D6')

for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

widths = {
    'A':12,'B':30,'C':32,'D':14,'E':58,'F':14,
    'G':18,'H':18,'I':18,'J':18,'K':16,'L':12,'M':20,
    'N':10,'O':10,'P':10,'Q':12,'R':16,'S':20,'T':20
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for r in range(2, ws.max_row + 1):
    ws[f'N{r}'].number_format = '0.0'
    ws[f'O{r}'].number_format = '0.0'
    ws[f'P{r}'].number_format = '0.0'
    ws[f'Q{r}'].number_format = '0.0%'
    estado = ws[f'F{r}'].value
    fill = other_fill
    if estado == 'TK_Complete':
        fill = complete_fill
    elif estado == 'TK_Active':
        fill = active_fill
    for c in range(1, 21):
        ws.cell(r, c).fill = fill

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:T{ws.max_row}"

for name in list(ws.tables.keys()):
    del ws.tables[name]

tab = Table(displayName='TablaRevisionAvanceW012v3', ref=f"A1:T{ws.max_row}")
style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

wb.save(out)
print(out)