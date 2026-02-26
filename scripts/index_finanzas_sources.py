import json
import zipfile
from pathlib import Path
from datetime import datetime
import openpyxl

BASE = Path(r"C:\Users\josej\OneDrive\002 SIMTEXX\000_Control de Porcesos\004_Informe M.Munizaga")
OUT = Path(r"C:\Users\josej\.openclaw\workspace\projects\finanzas\index")
OUT.mkdir(parents=True, exist_ok=True)

excel_index = []
pbix_index = []

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

for p in sorted(BASE.iterdir()):
    if not p.is_file():
        continue
    ext = p.suffix.lower()
    if ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(p, data_only=False)
            sheets = []
            for ws in wb.worksheets:
                formula_cells = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for c in row:
                        if isinstance(c.value, str) and c.value.startswith("="):
                            formula_cells += 1
                sheets.append({
                    "name": ws.title,
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                    "tables": list(ws.tables.keys()) if hasattr(ws, "tables") else [],
                    "formula_cells": formula_cells,
                })
            excel_index.append({
                "file": str(p),
                "name": p.name,
                "size": p.stat().st_size,
                "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
                "sheet_count": len(sheets),
                "sheets": sheets,
                "indexed_at": now(),
            })
        except Exception as e:
            excel_index.append({
                "file": str(p),
                "name": p.name,
                "error": str(e),
                "indexed_at": now(),
            })

    elif ext == ".pbix":
        entry = {
            "file": str(p),
            "name": p.name,
            "size": p.stat().st_size,
            "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
            "zip_entries": [],
            "has_datamodel": False,
            "has_layout": False,
            "has_connections": False,
            "indexed_at": now(),
        }
        try:
            with zipfile.ZipFile(p, "r") as z:
                names = z.namelist()
                entry["zip_entries"] = names[:200]
                lowered = [n.lower() for n in names]
                entry["has_datamodel"] = any("datamodel" in n for n in lowered)
                entry["has_layout"] = any("report/layout" in n for n in lowered)
                entry["has_connections"] = any("connections" in n for n in lowered)
        except Exception as e:
            entry["error"] = str(e)
        pbix_index.append(entry)

(OUT / "excel_index.json").write_text(json.dumps(excel_index, ensure_ascii=False, indent=2), encoding="utf-8")
(OUT / "pbix_index.json").write_text(json.dumps(pbix_index, ensure_ascii=False, indent=2), encoding="utf-8")

summary = {
    "base": str(BASE),
    "excel_files_indexed": len(excel_index),
    "pbix_files_indexed": len(pbix_index),
    "generated_at": now(),
    "outputs": [
        str(OUT / "excel_index.json"),
        str(OUT / "pbix_index.json"),
    ],
}
(OUT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(summary, ensure_ascii=False, indent=2))