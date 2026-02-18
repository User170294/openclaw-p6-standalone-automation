import argparse
import json
import re
from pathlib import Path
from openpyxl import load_workbook


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip()
        return t if t else None
    return v


def row_non_empty(row):
    return any(clean(x) is not None for x in row)


def looks_unit_row(row):
    text = " ".join(str(x).lower() for x in row if isinstance(x, str))
    return ("kgf/m" in text) or ("cm4" in text) or ("cm3" in text)


def normalize_headers(header_rows):
    # header_rows: list of rows (top->bottom), each row list
    max_cols = max(len(r) for r in header_rows)
    cols = []
    for c in range(max_cols):
        parts = []
        for r in header_rows:
            val = clean(r[c] if c < len(r) else None)
            if isinstance(val, str):
                if val not in parts:
                    parts.append(val)
        cols.append(" | ".join(parts) if parts else f"col_{c+1}")
    return cols


def extract_family(designation, sheet_name):
    if isinstance(designation, str):
        m = re.match(r"\s*([A-ZÑÁÉÍÓÚ¤]{1,4})\b", designation.upper())
        if m:
            return m.group(1)
    m2 = re.match(r"\s*([A-ZÑÁÉÍÓÚ]{1,4})\b", sheet_name.upper())
    return m2.group(1) if m2 else sheet_name.upper()


def to_number(x):
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip().replace(",", ".")
        try:
            return float(s)
        except Exception:
            return None
    return None


def process_sheet(ws):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    # Find likely unit/data boundary row
    unit_idx = None
    for i, r in enumerate(rows):
        if looks_unit_row(r):
            unit_idx = i
            break

    if unit_idx is None:
        return []

    header_start = max(0, unit_idx - 3)
    headers = normalize_headers(rows[header_start:unit_idx + 1])

    records = []
    current_designation = None
    for i in range(unit_idx + 1, len(rows)):
        r = rows[i]
        if not row_non_empty(r):
            continue

        vals = [clean(x) for x in r]

        # Designation usually in col2 (index 1)
        if len(vals) > 1 and isinstance(vals[1], str):
            current_designation = vals[1]
        designation = current_designation

        numeric_count = sum(1 for x in vals if isinstance(x, (int, float)))
        if numeric_count < 3:
            continue

        row_map = {}
        for c, v in enumerate(vals):
            if c < len(headers):
                key = headers[c]
            else:
                key = f"col_{c+1}"
            row_map[key] = v

        peso = to_number(vals[2] if len(vals) > 2 else None)

        rec = {
            "sheet": ws.title,
            "row": i + 1,
            "designation": designation,
            "family": extract_family(designation or "", ws.title),
            "peso_kgfm": peso,
            "values": row_map,
        }
        records.append(rec)

    return records


def main():
    ap = argparse.ArgumentParser(description="Construye índice JSONL desde catálogo ICHA (Excel)")
    ap.add_argument("--input", required=True, help="Ruta al .xlsx")
    ap.add_argument("--out-dir", default="data", help="Carpeta de salida")
    args = ap.parse_args()

    in_path = Path(args.input)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(in_path, data_only=True)

    all_records = []
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "indice":
            continue
        all_records.extend(process_sheet(ws))

    jsonl_path = out_dir / "icha_catalogo_normalizado.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as f:
        for rec in all_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # Simple index summary
    by_sheet = {}
    by_family = {}
    for r in all_records:
        by_sheet[r["sheet"]] = by_sheet.get(r["sheet"], 0) + 1
        by_family[r["family"]] = by_family.get(r["family"], 0) + 1

    idx = {
        "source": str(in_path),
        "records": len(all_records),
        "sheets": by_sheet,
        "families": by_family,
        "jsonl": str(jsonl_path),
    }

    idx_path = out_dir / "icha_indice.json"
    idx_path.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(idx, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
