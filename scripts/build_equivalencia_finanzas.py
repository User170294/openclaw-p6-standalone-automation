from openpyxl import load_workbook
from pathlib import Path
import json

excel_path = Path(r"C:\Users\josej\OneDrive\002 SIMTEXX\000_Control de Porcesos\004_Informe M.Munizaga\Presupuesto 2025 Simtexx Junio 2025.xlsx")
ws = load_workbook(excel_path, data_only=True, read_only=True)["RESUMEN"]

months = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
col0 = 4

rows = {
    "Gastos Adm y Ventas": 32,
    "Gastos Operaciones": 37,
    "Gastos Financieros": 50,
    "Total Gastos (calc)": None,
}

series = {}
for k,r in rows.items():
    if r is None:
        continue
    vals=[]
    for i in range(12):
        v=ws.cell(r,col0+i).value
        vals.append(float(v or 0))
    series[k]=vals
series["Total Gastos (calc)"]=[series["Gastos Adm y Ventas"][i]+series["Gastos Operaciones"][i]+series["Gastos Financieros"][i] for i in range(12)]

matrix = {
  "excel_source": str(excel_path),
  "sheet": "RESUMEN",
  "equivalencias": [
    {
      "excel_metrica": "Gastos Adm y Ventas (fila 32)",
      "pbix_equivalente": "No directo en visuales; probable componente dentro de Sum(001_Dat_Presupuesto.Valor) sujeto a filtros de tipo/categoría",
      "estado": "parcial"
    },
    {
      "excel_metrica": "Gastos Operaciones (fila 37)",
      "pbix_equivalente": "No directo en visuales; probable componente dentro de Sum(001_Dat_Presupuesto.Valor) sujeto a filtros",
      "estado": "parcial"
    },
    {
      "excel_metrica": "Gastos Financieros (fila 50)",
      "pbix_equivalente": "No explícito en visuales detectados",
      "estado": "no-evidencia"
    },
    {
      "excel_metrica": "Total Gastos mensual (32+37+50)",
      "pbix_equivalente": "No hay medida explícita de gastos totales; visual principal cruza Presupuesto vs Venta",
      "estado": "no-equivalente-directo"
    },
    {
      "excel_metrica": "Resultado Financiero Consolidado",
      "pbix_equivalente": "Visuales de página 1 usan Sum(001_Dat_Presupuesto.Valor) y Sum(002_Dat_Venta.Valor) por Mes",
      "estado": "equivalente-conceptual"
    }
  ],
  "totales_2025": {
    "Gastos Adm y Ventas": round(sum(series["Gastos Adm y Ventas"]),2),
    "Gastos Operaciones": round(sum(series["Gastos Operaciones"]),2),
    "Gastos Financieros": round(sum(series["Gastos Financieros"]),2),
    "Total Gastos (calc)": round(sum(series["Total Gastos (calc)"]),2)
  },
  "mensual_total_gastos": [
    {"mes": months[i], "total": round(series["Total Gastos (calc)"][i],2)} for i in range(12)
  ]
}

out = Path(r"C:\Users\josej\.openclaw\workspace\projects\finanzas\index\equivalencia_excel_pbix_gastos_2025.json")
out.write_text(json.dumps(matrix, ensure_ascii=False, indent=2), encoding="utf-8")
print(out)
print(json.dumps(matrix["totales_2025"], ensure_ascii=False))
