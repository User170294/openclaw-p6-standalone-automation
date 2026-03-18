"""Tests for prepare_load_workbook_from_capture.py — workbook preparation logic."""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_CORE = ROOT / "scripts" / "core"
if str(SCRIPTS_CORE) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_CORE))

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook as lw
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_capture_xlsx(tmp_path: Path, sheet_name: str, task_codes: list[str]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ["Entrega", "Despacho/Tag", "WBS", "Actividad ID", "Actividad",
               "Estado P6", "BAC HH", "EV HH", "Avance %", "Marcar avance"]
    ws.append(headers)
    for code in task_codes:
        ws.append(["ENT-1", "TAG-1", "WBS-1", code, f"Tarea {code}", "TK_Active",
                   100, 50, 0.5, ""])
    path = tmp_path / "capture.xlsx"
    wb.save(path)
    return path


def _make_db(tmp_path: Path, tasks: list[dict]) -> Path:
    db_path = tmp_path / "test.db"
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.executescript("""
        CREATE TABLE PROJECT (
            PROJ_ID INTEGER PRIMARY KEY,
            PROJ_SHORT_NAME TEXT,
            DEF_COMPLETE_PCT_TYPE TEXT DEFAULT 'CP_Drtn'
        );
        CREATE TABLE TASK (
            TASK_ID INTEGER PRIMARY KEY,
            PROJ_ID INTEGER,
            TASK_CODE TEXT,
            TASK_NAME TEXT,
            COMPLETE_PCT_TYPE TEXT,
            CLNDR_ID INTEGER
        );
    """)
    cur.execute("INSERT INTO PROJECT VALUES (1, 'TEST', 'CP_Drtn')")
    for t in tasks:
        cur.execute("INSERT INTO TASK VALUES (?,1,?,?,?,?)",
                    (t["id"], t["code"], t["name"], t.get("pct_type"), t.get("clndr_id")))
    con.commit()
    con.close()
    return db_path


# ---------------------------------------------------------------------------
# main() integration
# ---------------------------------------------------------------------------

def test_main_adds_complete_pct_type_and_calendar_columns(tmp_path):
    db = _make_db(tmp_path, [
        {"id": 1, "code": "A001", "name": "Tarea A", "pct_type": "CP_Drtn", "clndr_id": 7475},
    ])
    sheet = "Revision_Avance_2026-W12"
    xlsx = _make_capture_xlsx(tmp_path, sheet, ["A001"])
    out = tmp_path / "out.xlsx"

    sys.argv = [
        "prepare_load_workbook_from_capture.py",
        "--db", str(db),
        "--source-xlsx", str(xlsx),
        "--out-xlsx", str(out),
        "--proj-id", "1",
        "--sheet", sheet,
    ]

    import prepare_load_workbook_from_capture as plw
    plw.main()

    assert out.exists()
    wb = lw(out)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    assert "% Complete Type" in headers
    assert "Calendar ID" in headers


def test_main_populates_complete_pct_type_from_db(tmp_path):
    db = _make_db(tmp_path, [
        {"id": 1, "code": "A001", "name": "Tarea A", "pct_type": "CP_Drtn", "clndr_id": 7475},
    ])
    sheet = "Revision_Avance_2026-W12"
    xlsx = _make_capture_xlsx(tmp_path, sheet, ["A001"])
    out = tmp_path / "out.xlsx"

    sys.argv = [
        "prepare_load_workbook_from_capture.py",
        "--db", str(db),
        "--source-xlsx", str(xlsx),
        "--out-xlsx", str(out),
        "--proj-id", "1",
        "--sheet", sheet,
    ]

    import prepare_load_workbook_from_capture as plw
    plw.main()

    wb = lw(out)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    pct_col = headers.index("% Complete Type") + 1
    cal_col = headers.index("Calendar ID") + 1

    # Row 2 = first data row
    assert ws.cell(2, pct_col).value == "CP_Drtn"
    assert ws.cell(2, cal_col).value == 7475


def test_main_raises_on_missing_sheet(tmp_path):
    db = _make_db(tmp_path, [])
    xlsx = _make_capture_xlsx(tmp_path, "Hoja_Correcta", [])
    out = tmp_path / "out.xlsx"

    sys.argv = [
        "prepare_load_workbook_from_capture.py",
        "--db", str(db),
        "--source-xlsx", str(xlsx),
        "--out-xlsx", str(out),
        "--proj-id", "1",
        "--sheet", "Hoja_Inexistente",
    ]

    import prepare_load_workbook_from_capture as plw
    with pytest.raises(SystemExit):
        plw.main()


def test_main_skips_rows_without_task_code(tmp_path):
    db = _make_db(tmp_path, [
        {"id": 1, "code": "A001", "name": "Tarea A", "pct_type": "CP_Drtn", "clndr_id": 7475},
    ])
    sheet = "Revision_Avance_2026-W12"

    # xlsx con una fila sin task_code
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Entrega", "Despacho/Tag", "WBS", "Actividad ID", "Actividad",
               "Estado P6", "BAC HH", "EV HH", "Avance %", "Marcar avance"])
    ws.append(["ENT-1", "TAG-1", "WBS-1", None, "Sin código", "TK_Active", 0, 0, 0, ""])
    xlsx = tmp_path / "capture.xlsx"
    wb.save(xlsx)
    out = tmp_path / "out.xlsx"

    sys.argv = [
        "prepare_load_workbook_from_capture.py",
        "--db", str(db),
        "--source-xlsx", str(str(xlsx)),
        "--out-xlsx", str(out),
        "--proj-id", "1",
        "--sheet", sheet,
    ]

    import prepare_load_workbook_from_capture as plw
    plw.main()

    assert out.exists()
    result_wb = lw(out)
    result_ws = result_wb[sheet]
    headers = [c.value for c in result_ws[1]]
    pct_col = headers.index("% Complete Type") + 1
    # Fila sin task_code debe tener None en % Complete Type
    assert result_ws.cell(2, pct_col).value is None
