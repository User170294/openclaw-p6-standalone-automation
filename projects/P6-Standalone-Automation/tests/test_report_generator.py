from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_CORE = ROOT / 'scripts' / 'core'
if str(SCRIPTS_CORE) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_CORE))

import report_generator  # noqa: E402


SAMPLE_REPORT = {
    'mode': 'logic',
    'stub': False,
    'note': 'Reporte de prueba para render final.',
    'row_count': 2,
    'columns': report_generator.FIXED_COLUMNS,
    'bac': 300.0,
    'pv': 300.0,
    'ev': 250.0,
    'ac': 260.0,
    'sv': -50.0,
    'cv': -10.0,
    'spi': 0.833333,
    'cpi': 0.961538,
    'eac': 312.0,
    'etc': 62.0,
    'rows': [
        {
            'week': 'Y26-W08', 'pv_week': 100.0, 'pv_cum': 100.0, 'pv_pct': 33.33,
            'ev_cum': 80.0, 'ev_pct': 26.67, 'ac_cum': 85.0,
            'sv': -20.0, 'spi': 0.8,
            'forecast_cum': 220.0, 'forecast_pct': 73.33,
        },
        {
            'week': 'Y26-W09', 'pv_week': 200.0, 'pv_cum': 300.0, 'pv_pct': 100.0,
            'ev_cum': 250.0, 'ev_pct': 83.33, 'ac_cum': 260.0,
            'sv': -50.0, 'spi': 0.833333,
            'forecast_cum': 300.0, 'forecast_pct': 100.0,
        },
    ],
}

EXPECTED_COLUMNS = ['week', 'pv_week', 'pv_cum', 'pv_pct', 'ev_cum', 'ev_pct', 'ac_cum', 'sv', 'spi', 'forecast_cum', 'forecast_pct']


def test_render_md_generates_non_empty_file(tmp_path):
    out = report_generator.generate_report(SAMPLE_REPORT, tmp_path, 'md', {'project_name': 'Proyecto Test'})
    text = out.read_text(encoding='utf-8')
    assert out.exists()
    assert out.stat().st_size > 0
    assert '# Proyecto Test' in text
    # Verifica que las columnas nuevas están en la tabla
    assert 'ac_cum' in text
    assert 'forecast_cum' in text
    # Verifica KPIs completos
    assert 'CPI' in text
    assert 'EAC' in text
    assert 'ETC' in text
    assert 'CV' in text
    assert 'AC' in text


def test_render_rejects_removed_html_format(tmp_path):
    try:
        report_generator.generate_report(SAMPLE_REPORT, tmp_path, 'html', {'project_name': 'Test'})
    except ValueError as exc:
        assert 'Formato de reporte no soportado' in str(exc)
    else:
        raise AssertionError('html should not be supported anymore')


def test_render_xlsx_generates_expected_sheets_and_columns(tmp_path):
    out = report_generator.generate_report(SAMPLE_REPORT, tmp_path, 'xlsx', {'project_name': 'Proyecto Test'})
    assert out.exists()
    assert out.stat().st_size > 0
    wb = load_workbook(out)
    assert 'weekly_data' in wb.sheetnames
    assert 'summary_chart' in wb.sheetnames

    ws = wb['weekly_data']
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(EXPECTED_COLUMNS) + 1)]
    assert headers == EXPECTED_COLUMNS

    # Verifica KPIs en summary_chart
    summary = wb['summary_chart']
    kpi_labels = [summary.cell(row=i, column=1).value for i in range(3, 13)]
    assert 'BAC' in kpi_labels
    assert 'EV' in kpi_labels
    assert 'AC' in kpi_labels
    assert 'CPI' in kpi_labels
    assert 'EAC' in kpi_labels
    assert 'ETC' in kpi_labels
    assert 'CV' in kpi_labels
