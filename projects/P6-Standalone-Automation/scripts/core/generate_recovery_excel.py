"""
generate_recovery_excel.py
--------------------------
Genera el Excel de Plan de Recovery semanal para un programa P6 SQLite.
Agrupa actividades por TAG > Entregable (WBS), calcula HH de la semana
objetivo usando el metodo de calendario real (CLNDR_DATA + horas efectivas).

Uso:
    python generate_recovery_excel.py
        --db  <ruta_db>
        --proj-id <PROJ_ID>
        --cutoff <YYYY-MM-DD>
        --week-start <YYYY-MM-DD>   # lunes de la semana objetivo (default: cutoff + 1 dia)
        --week-end   <YYYY-MM-DD>   # domingo (default: week-start + 6)
        --bac        <float>        # BAC total HH (default: sum TARGET_QTY RT_Labor)
        --ev         <float>        # EV acumulado HH (default: sum ACT_REG_QTY RT_Labor)
        --pv-week    <float>        # PV semana HH (default: 0, puede pasarse manualmente)
        --out        <ruta_xlsx>    # destino (default: <proj_id>_Recovery_<week>.xlsx)

Requiere: openpyxl
"""
from __future__ import annotations

import argparse
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")


# ---------------------------------------------------------------------------
# Calendario helpers (misma logica que pv_engine.py)
# ---------------------------------------------------------------------------

def _excel_serial_to_date(serial: int) -> date:
    return (datetime(1899, 12, 30) + timedelta(days=int(serial))).date()


def parse_clock(value: str):
    return datetime.strptime(value, "%H:%M").time()


def combine_dt(day: date, clock) -> datetime:
    return datetime(day.year, day.month, day.day, clock.hour, clock.minute)


def overlap_hours(a0: datetime, a1: datetime, b0: datetime, b1: datetime) -> float:
    s = max(a0, b0)
    e = min(a1, b1)
    return max(0.0, (e - s).total_seconds() / 3600.0)


def parse_calendar_text(raw: Any) -> dict:
    """Devuelve {weekday_windows: {0-6: [(start, end)]}, exceptions: {date}}"""
    text = "" if raw is None else str(raw)
    if text.startswith("b'") and text.endswith("'"):
        text = (
            text[2:-1]
            .encode("latin-1", errors="ignore")
            .decode("unicode_escape", errors="ignore")
        )
    weekday_windows: dict[int, list] = defaultdict(list)
    exceptions: set[date] = set()
    days_start = text.find("DaysOfWeek()(")
    view_start = text.find("(0||VIEW", days_start if days_start >= 0 else 0)
    days_text = (
        text[days_start:view_start]
        if days_start >= 0 and view_start > days_start
        else text
    )
    markers = list(re.finditer(r"\(0\|\|([1-7])\(\)", days_text))
    for idx, match in enumerate(markers):
        p6_day = int(match.group(1))
        body_start = match.end()
        body_end = markers[idx + 1].start() if idx + 1 < len(markers) else len(days_text)
        body = days_text[body_start:body_end]
        py_weekday = (p6_day + 5) % 7
        for s_txt, f_txt in re.findall(r"s\|(\d{2}:\d{2})\|f\|(\d{2}:\d{2})", body):
            weekday_windows[py_weekday].append(
                (parse_clock(s_txt), parse_clock(f_txt))
            )
    for serial_txt in re.findall(r"d\|(\d+)", text):
        exceptions.add(_excel_serial_to_date(int(serial_txt)))
    for wd in list(weekday_windows):
        weekday_windows[wd] = sorted(weekday_windows[wd], key=lambda x: x[0])
    return {"weekday_windows": dict(weekday_windows), "exceptions": exceptions}


def week_hh(
    remain_start: datetime,
    end_dt: datetime,
    qty: float,
    calendar: dict | None,
    week_mon: date,
    week_sun: date,
) -> float:
    """Fraccion de qty que cae dentro de [week_mon, week_sun] usando horas efectivas."""
    if not end_dt or end_dt < remain_start or qty <= 0:
        return 0.0
    weekday_windows = (calendar or {}).get("weekday_windows") or {}
    exceptions = (calendar or {}).get("exceptions") or set()
    hours_all: dict[date, float] = {}
    hours_week: dict[date, float] = {}
    d = remain_start.date()
    while d <= end_dt.date():
        if d not in exceptions:
            windows = weekday_windows.get(d.weekday(), [])
            hrs = sum(
                overlap_hours(remain_start, end_dt, combine_dt(d, ws), combine_dt(d, we))
                for ws, we in windows
            )
            if hrs > 0:
                hours_all[d] = hrs
                if week_mon <= d <= week_sun:
                    hours_week[d] = hrs
        d += timedelta(days=1)
    total = sum(hours_all.values())
    if total <= 0:
        return 0.0
    return qty * sum(hours_week.values()) / total


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def safe_str(v: Any) -> str:
    if v is None:
        return ""
    try:
        return str(v).encode("latin-1", errors="replace").decode("utf-8", errors="replace")
    except Exception:
        return str(v)


def safe_dt(v: Any) -> datetime | None:
    if not v:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v), fmt)
        except Exception:
            pass
    return None


def load_calendars(cur: sqlite3.Cursor, proj_id: int) -> dict[int, dict]:
    rows = cur.execute(
        """SELECT c.CLNDR_ID, c.CLNDR_DATA FROM CALENDAR c
           WHERE c.CLNDR_ID IN (
               SELECT DISTINCT CLNDR_ID FROM TASK
               WHERE PROJ_ID=? AND CLNDR_ID IS NOT NULL)""",
        (proj_id,),
    ).fetchall()
    return {int(r["CLNDR_ID"]): parse_calendar_text(r["CLNDR_DATA"]) for r in rows}


def load_wbs(cur: sqlite3.Cursor, proj_id: int) -> tuple[dict, dict, int | None]:
    """Devuelve (wbs_map, tag_map, entregable_map, fab_parent_id)."""
    rows = cur.execute(
        "SELECT WBS_ID, PARENT_WBS_ID, WBS_SHORT_NAME, WBS_NAME FROM PROJWBS WHERE PROJ_ID=?",
        (proj_id,),
    ).fetchall()
    wbs_map: dict[int, dict] = {}
    for r in rows:
        wbs_map[r["WBS_ID"]] = {
            "name": safe_str(r["WBS_NAME"]),
            "short": safe_str(r["WBS_SHORT_NAME"]),
            "parent": r["PARENT_WBS_ID"],
        }

    # Detectar nodo raíz de fabricación (nivel 3 típico con múltiples entregables)
    # Heuristica: el nodo cuyo short empieza con numero y tiene >2 hijos
    fab_parent_id: int | None = None
    children_count: dict[int, int] = defaultdict(int)
    for w in wbs_map.values():
        if w["parent"]:
            children_count[w["parent"]] += 1
    # Buscar nodo con short numerico y muchos hijos
    for wid, w in wbs_map.items():
        if children_count.get(wid, 0) >= 3:
            # verificar que sus hijos tambien tienen hijos (nivel TAG)
            child_ids = [k for k, v in wbs_map.items() if v["parent"] == wid]
            if any(children_count.get(c, 0) >= 2 for c in child_ids):
                fab_parent_id = wid
                break

    tag_map: dict[int, str] = {}
    entregable_map: dict[int, str] = {}
    if fab_parent_id:
        for wid, w in wbs_map.items():
            if w["parent"] == fab_parent_id:
                # Extraer numero de TAG del nombre
                m = re.search(r"-(\d+)$", w["name"])
                tag_label = f"TAG-{m.group(1)}" if m else w["short"]
                tag_map[wid] = tag_label
                for wid2, w2 in wbs_map.items():
                    if w2["parent"] == wid:
                        raw_name = w2["name"]
                        clean = re.sub(r"\s*\(.*?\)", "", raw_name).strip()
                        entregable_map[wid2] = clean

    return wbs_map, tag_map, entregable_map


def get_tag_entregable(
    wbs_id: int,
    wbs_map: dict,
    tag_map: dict,
    entregable_map: dict,
) -> tuple[str, str]:
    cur_id = wbs_id
    for _ in range(8):
        if cur_id in entregable_map:
            node = wbs_map.get(cur_id, {})
            tag = tag_map.get(node.get("parent"), "")
            return tag, entregable_map[cur_id]
        if cur_id in tag_map:
            return tag_map[cur_id], ""
        node = wbs_map.get(cur_id)
        if not node:
            break
        cur_id = node["parent"]
    return "", ""


def load_activities(
    cur: sqlite3.Cursor,
    proj_id: int,
    cutoff: datetime,
    week_mon: date,
    week_sun: date,
    calendars: dict,
    wbs_map: dict,
    tag_map: dict,
    entregable_map: dict,
) -> list[dict]:
    rows = cur.execute(
        """
        SELECT t.TASK_CODE, t.TASK_NAME, t.STATUS_CODE, t.CLNDR_ID,
               t.EARLY_START_DATE, t.EARLY_END_DATE, t.WBS_ID,
               COALESCE(SUM(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN tr.REMAIN_QTY ELSE 0 END),0) AS remain_qty
        FROM TASK t
        LEFT JOIN TASKRSRC tr ON tr.PROJ_ID=t.PROJ_ID AND tr.TASK_ID=t.TASK_ID
        WHERE t.PROJ_ID=? AND t.TASK_TYPE != 'TT_Mile'
        GROUP BY t.TASK_ID, t.TASK_CODE, t.TASK_NAME, t.STATUS_CODE,
                 t.CLNDR_ID, t.EARLY_START_DATE, t.EARLY_END_DATE, t.WBS_ID
        HAVING remain_qty > 0
        """,
        (proj_id,),
    ).fetchall()

    STATUS_LABEL = {
        "TK_NotStart": "No Iniciada",
        "TK_Active": "En Curso",
        "TK_Complete": "Completada",
    }

    result = []
    for r in rows:
        remain = float(r["remain_qty"] or 0)
        early_start = safe_dt(r["EARLY_START_DATE"])
        early_end = safe_dt(r["EARLY_END_DATE"])
        if not early_end:
            continue
        start_dt = cutoff + timedelta(seconds=1)
        if early_start and early_start > start_dt:
            start_dt = early_start
        cal = calendars.get(int(r["CLNDR_ID"])) if r["CLNDR_ID"] is not None else None
        hh = week_hh(start_dt, early_end, remain, cal, week_mon, week_sun)
        if hh <= 0:
            continue
        tag, entregable = get_tag_entregable(r["WBS_ID"], wbs_map, tag_map, entregable_map)
        result.append(
            {
                "tag": tag or "Sin TAG",
                "entregable": entregable or "Sin Entregable",
                "task_code": safe_str(r["TASK_CODE"]),
                "task_name": safe_str(r["TASK_NAME"]),
                "status": STATUS_LABEL.get(safe_str(r["STATUS_CODE"]), safe_str(r["STATUS_CODE"])),
                "early_start": str(r["EARLY_START_DATE"])[:16] if r["EARLY_START_DATE"] else "",
                "early_end": str(r["EARLY_END_DATE"])[:16] if r["EARLY_END_DATE"] else "",
                "remain_hh": round(remain, 1),
                "w_hh": round(hh, 1),
            }
        )
    return sorted(result, key=lambda x: (x["tag"], x["entregable"], -x["w_hh"], x["task_code"]))


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    t = Side(border_style="thin", color="BDC3C7")
    return Border(left=t, right=t, top=t, bottom=t)


PALETTE = {
    "header_dark": "1F3864",
    "header_mid": "2E75B6",
    "tag_bg": "D6E4F0",
    "tag_fg": "1F3864",
    "ent_bg": "EBF5FB",
    "ent_fg": "154360",
    "kpi_bg": "D6E4F0",
    "footer_bg": "1F3864",
    "odd": "F8FBFF",
    "even": "FFFFFF",
    "green": "D5F5E3",
    "orange": "FDEBD0",
    "blue_light": "D6EAF8",
    "hh_high": "1E8449",
    "hh_mid": "1F3864",
    "hh_low": "C0392B",
}


def build_excel(
    activities: list[dict],
    out_path: Path,
    proj_id: int,
    week_mon: date,
    week_sun: date,
    bac: float,
    ev: float,
    pv_week: float,
    project_name: str = "",
) -> None:
    if not project_name:
        project_name = f"PROJ_ID={proj_id}"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recovery"
    BORDER = _thin_border()

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def font(color: str = "000000", bold: bool = False, size: int = 9) -> Font:
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def cell(row: int, col: int, value: Any = None) -> Any:
        c = ws.cell(row, col)
        if value is not None:
            c.value = value
        c.border = BORDER
        return c

    def merge_row(row: int, col_start: int, col_end: int, value: str, f: Font, p: PatternFill, align: str = "left") -> None:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row, col_start)
        c.value = value
        c.font = f
        c.fill = p
        c.border = BORDER
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        for col in range(col_start + 1, col_end + 1):
            ws.cell(row, col).border = BORDER

    NCOLS = 9
    iso = week_mon.isocalendar()
    week_label = f"Y{iso.year % 100:02d}-W{iso.week:02d}"
    re_week = sum(a["w_hh"] for a in activities)
    fc = ev + re_week
    fc_pct = fc / bac * 100 if bac else 0
    pv_cum_w = 0.0  # placeholder si no viene calculado

    # Row 1: titulo
    ws.row_dimensions[1].height = 22
    merge_row(
        1, 1, NCOLS,
        f"PLAN DE RECOVERY {week_label}  |  {project_name}  |  {week_mon.strftime('%d-%m-%Y')} a {week_sun.strftime('%d-%m-%Y')}",
        Font(bold=True, color="FFFFFF", size=12, name="Calibri"),
        fill("1A252F"), "center",
    )

    # Row 2: KPI bar
    ws.row_dimensions[2].height = 16
    kpi_txt = (
        f"BAC: {bac:,.1f} HH   |   "
        f"EV acum: {ev:,.1f} HH ({ev/bac*100:.2f}%)   |   "
        f"RE {week_label}: {re_week:,.1f} HH   |   "
        f"Forecast fin {week_label}: {fc:,.1f} HH ({fc_pct:.2f}%)"
    )
    merge_row(2, 1, NCOLS, kpi_txt, font(PALETTE["tag_fg"], bold=True, size=9), fill(PALETTE["kpi_bg"]), "center")

    # Row 3: headers
    HEADERS = ["TAG", "Entregable", "Actividad", "Nombre Actividad", "Estado", "Early Start", "Early End", "Remain HH", f"{week_label} HH"]
    WIDTHS =  [12,    28,           10,           44,                 12,       16,             16,           10,           10]
    ws.row_dimensions[3].height = 20
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = cell(3, i, h)
        c.font = font("FFFFFF", bold=True, size=10)
        c.fill = fill(PALETTE["header_dark"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    STATUS_FILL = {
        "En Curso":   PALETTE["green"],
        "No Iniciada": PALETTE["orange"],
        "Completada": PALETTE["blue_light"],
    }
    row_num = 4
    prev_tag = None
    prev_ent = None
    row_color = 0

    for rec in activities:
        if rec["tag"] != prev_tag:
            ws.row_dimensions[row_num].height = 14
            merge_row(
                row_num, 1, NCOLS,
                f"\u25B6  {rec['tag']}",
                font(PALETTE["tag_fg"], bold=True, size=9),
                fill(PALETTE["tag_bg"]), "left",
            )
            row_num += 1
            prev_tag = rec["tag"]
            prev_ent = None
            row_color = 0

        if rec["entregable"] != prev_ent:
            ws.row_dimensions[row_num].height = 13
            ws.cell(row_num, 1).fill = fill(PALETTE["ent_bg"])
            ws.cell(row_num, 1).border = BORDER
            merge_row(
                row_num, 2, NCOLS,
                f"    {rec['entregable']}",
                font(PALETTE["ent_fg"], bold=True, size=9),
                fill(PALETTE["ent_bg"]), "left",
            )
            row_num += 1
            prev_ent = rec["entregable"]
            row_color = 0

        bg = PALETTE["odd"] if row_color % 2 == 0 else PALETTE["even"]
        row_color += 1
        ws.row_dimensions[row_num].height = 13
        vals = [
            rec["tag"], rec["entregable"], rec["task_code"], rec["task_name"],
            rec["status"], rec["early_start"], rec["early_end"],
            rec["remain_hh"], rec["w_hh"],
        ]
        for i, v in enumerate(vals, 1):
            c = cell(row_num, i, v)
            c.fill = fill(STATUS_FILL[rec["status"]]) if i == 5 else fill(bg)
            if i in (8, 9):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center" if i in (1, 3) else "left", vertical="center", wrap_text=True)

        # W_HH color by magnitude
        hh_val = rec["w_hh"]
        hh_color = PALETTE["hh_high"] if hh_val >= 54 else (PALETTE["hh_low"] if hh_val < 18 else PALETTE["hh_mid"])
        ws.cell(row_num, 9).font = font(hh_color, bold=True, size=9)
        row_num += 1

    # Footer totals
    ws.row_dimensions[row_num].height = 16
    for i in range(1, NCOLS + 1):
        c = cell(row_num, i)
        c.fill = fill(PALETTE["footer_bg"])
        c.font = font("FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    ws.cell(row_num, 1).value = f"TOTAL {week_label}   —   {len(activities)} actividades"
    ws.cell(row_num, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row_num, 8).value = round(sum(a["remain_hh"] for a in activities), 1)
    ws.cell(row_num, 9).value = round(re_week, 1)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Recovery Excel semanal desde P6 SQLite")
    p.add_argument("--db", required=True)
    p.add_argument("--proj-id", type=int, required=True)
    p.add_argument("--cutoff", required=True, help="YYYY-MM-DD")
    p.add_argument("--week-start", default=None, help="YYYY-MM-DD lunes semana objetivo")
    p.add_argument("--week-end", default=None, help="YYYY-MM-DD domingo semana objetivo")
    p.add_argument("--bac", type=float, default=None)
    p.add_argument("--ev", type=float, default=None)
    p.add_argument("--pv-week", type=float, default=0.0)
    p.add_argument("--project-name", default="", help="Nombre del proyecto para encabezado del reporte (default: PROJ_ID=<id>)")
    p.add_argument("--out", default=None)
    return p.parse_args()


def main() -> None:
    args = _parse_args()
    cutoff = datetime.strptime(args.cutoff, "%Y-%m-%d").replace(hour=23, minute=59, second=0)
    week_mon = (
        datetime.strptime(args.week_start, "%Y-%m-%d").date()
        if args.week_start
        else (cutoff + timedelta(days=1)).date()
    )
    week_sun = (
        datetime.strptime(args.week_end, "%Y-%m-%d").date()
        if args.week_end
        else week_mon + timedelta(days=6)
    )

    con = sqlite3.connect(args.db)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    calendars = load_calendars(cur, args.proj_id)
    wbs_map, tag_map, entregable_map = load_wbs(cur, args.proj_id)

    # BAC y EV desde DB si no se pasan
    bac = args.bac
    ev = args.ev
    if bac is None:
        row = cur.execute(
            "SELECT COALESCE(SUM(TARGET_QTY),0) AS v FROM TASKRSRC WHERE PROJ_ID=? AND RSRC_TYPE='RT_Labor'",
            (args.proj_id,),
        ).fetchone()
        bac = float(row["v"] or 0)
    if ev is None:
        row = cur.execute(
            "SELECT COALESCE(SUM(ACT_REG_QTY),0) AS v FROM TASKRSRC WHERE PROJ_ID=? AND RSRC_TYPE='RT_Labor'",
            (args.proj_id,),
        ).fetchone()
        ev = float(row["v"] or 0)

    activities = load_activities(cur, args.proj_id, cutoff, week_mon, week_sun, calendars, wbs_map, tag_map, entregable_map)
    con.close()

    re_week = sum(a["w_hh"] for a in activities)
    fc = ev + re_week
    iso = week_mon.isocalendar()
    week_label = f"Y{iso.year % 100:02d}-W{iso.week:02d}"

    out_path = Path(args.out) if args.out else Path(f"{args.proj_id}_Recovery_{week_label}.xlsx")
    build_excel(activities, out_path, args.proj_id, week_mon, week_sun, bac, ev, args.pv_week, project_name=args.project_name)

    print(f"PROJ_ID   : {args.proj_id}")
    print(f"Cutoff    : {cutoff.date()}")
    print(f"Semana    : {week_label}  ({week_mon} — {week_sun})")
    print(f"BAC       : {bac:,.1f} HH")
    print(f"EV acum   : {ev:,.1f} HH  ({ev/bac*100:.2f}%)" if bac else f"EV        : {ev:,.1f} HH")
    print(f"RE {week_label}   : {re_week:,.1f} HH")
    print(f"Forecast  : {fc:,.1f} HH  ({fc/bac*100:.2f}%)" if bac else f"Forecast  : {fc:,.1f} HH")
    print(f"Actividades: {len(activities)}")
    print(f"OUT       : {out_path}")


if __name__ == "__main__":
    main()
