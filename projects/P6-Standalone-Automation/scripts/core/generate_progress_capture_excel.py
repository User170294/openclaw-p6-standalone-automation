#!/usr/bin/env python3
import argparse
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPTS_ROOT = Path(__file__).resolve().parents[1]
if str(SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_ROOT))

from p6_utils import open_db

SHEET_NAME = 'Revision_Avance_W012'
HEADERS = [
    'Actividad ID',
    'Actividad',
    'Estado P6',
    'Entrega',
    'Paquete',
    'Ruta WBS',
    '% Complete Type',
    'Calendar ID',
    'Inicio plan',
    'TÃƒÂ©rmino plan',
    'BAC HH',
    'EV HH',
    'ETC HH',
    '% avance a cargar',
    'Fecha inicio real a cargar',
    'Fecha tÃƒÂ©rmino si 100%',
    'Notas usuario',
]

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(color='FFFFFF', bold=True)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP_ALIGN = Alignment(vertical='top', wrap_text=True)

DELIVERY_FILLS = [
    PatternFill('solid', fgColor='FDE9D9'),
    PatternFill('solid', fgColor='EAF2D3'),
    PatternFill('solid', fgColor='DDEBF7'),
    PatternFill('solid', fgColor='EADCF8'),
    PatternFill('solid', fgColor='FFF2CC'),
]


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description='Genera Excel de captura de avances semanales desde DB SQLite de P6.')
    ap.add_argument('--db', required=True)
    ap.add_argument('--proj-id', type=int, required=True)
    ap.add_argument('--iso-week', required=True, help='Formato ISO YYYY-W##, ej. 2026-W11')
    ap.add_argument('--sheet', default=SHEET_NAME)
    ap.add_argument('--out', required=True, help='Ruta xlsx de salida')
    return ap.parse_args()


def iso_week_bounds(iso_week: str) -> tuple[date, date]:
    txt = iso_week.strip().upper().replace('ISO', '')
    year_s, week_s = txt.split('-W')
    year = int(year_s)
    week = int(week_s)
    start = date.fromisocalendar(year, week, 1)
    end = date.fromisocalendar(year, week, 7)
    return start, end


def parse_db_dt(value: Any) -> datetime | None:
    if value in (None, ''):
        return None
    txt = str(value)
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f'):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            pass
    raise ValueError(f'Fecha DB no reconocida: {value!r}')


def overlaps(start_a: datetime | None, end_a: datetime | None, start_b: datetime, end_b: datetime) -> bool:
    if start_a is None and end_a is None:
        return False
    left = start_a or end_a
    right = end_a or start_a
    if left is None or right is None:
        return False
    return left <= end_b and right >= start_b


def auto_fit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            size = len(str(value)) if value is not None else 0
            widths[idx] = max(widths.get(idx, 0), min(size + 2, 90))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_header(ws, headers: list[str]) -> None:
    ws.row_dimensions[1].height = 32
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def clean_label(short_name: str | None, name: str | None) -> str:
    short = (short_name or '').strip()
    name = (name or '').strip()
    if short and name and short != name:
        return f'{short} | {name}'
    return name or short


def build_wbs_maps(cur: sqlite3.Cursor, proj_id: int) -> tuple[dict[int, sqlite3.Row], dict[int, list[str]]]:
    rows = cur.execute(
        '''
        SELECT WBS_ID, PARENT_WBS_ID, WBS_SHORT_NAME, WBS_NAME
        FROM PROJWBS
        WHERE PROJ_ID = ?
        ''',
        (proj_id,),
    ).fetchall()
    by_id = {int(r['WBS_ID']): r for r in rows}
    chain_map: dict[int, list[str]] = {}

    def compose(wbs_id: int | None) -> list[str]:
        if not wbs_id or wbs_id not in by_id:
            return []
        if wbs_id in chain_map:
            return chain_map[wbs_id]
        chain_rows: list[sqlite3.Row] = []
        current = by_id[wbs_id]
        seen: set[int] = set()
        while current and int(current['WBS_ID']) not in seen:
            seen.add(int(current['WBS_ID']))
            chain_rows.append(current)
            parent_id = current['PARENT_WBS_ID']
            current = by_id.get(int(parent_id)) if parent_id is not None and int(parent_id) in by_id else None
        chain_rows.reverse()
        labels = [clean_label(r['WBS_SHORT_NAME'], r['WBS_NAME']) for r in chain_rows if clean_label(r['WBS_SHORT_NAME'], r['WBS_NAME'])]
        chain_map[wbs_id] = labels
        return labels

    for wbs_id in by_id:
        compose(wbs_id)
    return by_id, chain_map


def derive_structure(parts: list[str]) -> tuple[str, str, str]:
    # Esperado OT-1844 W12:
    # 0 programa, 1 macro (FabricaciÃƒÂ³n 10 sistemas...), 2 entrega TAG..., 3 paquete (Bastidor, etc)
    entrega = parts[2] if len(parts) >= 3 else (parts[-1] if parts else '')
    paquete = parts[3] if len(parts) >= 4 else ''
    ruta = ' > '.join(parts)
    return entrega, paquete, ruta


def apply_delivery_fill(ws, row_idx: int, entrega: str, delivery_palette: dict[str, PatternFill]) -> None:
    if not entrega:
        return
    fill = delivery_palette.get(entrega)
    if not fill:
        return
    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet('Instrucciones')
    rows = [
        ('Objetivo', 'Esta hoja sirve para capturar el avance real de W11 con una vista clara por entrega y paquete.'),
        ('CÃƒÂ³mo leer la estructura', 'La columna Entrega separa las 5 entregas. La columna Paquete indica el subconjunto dentro de cada entrega (Bastidor, Bandeja, Barandas, etc). Ruta WBS muestra la jerarquÃƒÂ­a completa.'),
        ('QuÃƒÂ© columnas editar', 'Solo editar: % avance a cargar, Fecha inicio real a cargar, Fecha tÃƒÂ©rmino si 100%, Notas usuario.'),
        ('QuÃƒÂ© no tocar', 'No modificar Actividad ID, Actividad, Estado P6, Entrega, Paquete, Ruta WBS, % Complete Type, Calendar ID, Inicio/TÃƒÂ©rmino plan, BAC HH, EV HH, ETC HH.'),
        ('Avance parcial', 'Si la actividad trabajÃƒÂ³ pero no terminÃƒÂ³: ingresar % avance a cargar y Fecha inicio real a cargar. Dejar vacÃƒÂ­a Fecha tÃƒÂ©rmino si 100%.'),
        ('Cierre 100%', 'Si la actividad terminÃƒÂ³: ingresar 100 en % avance a cargar, Fecha inicio real a cargar y Fecha tÃƒÂ©rmino si 100%.'),
        ('Uso recomendado', 'Filtra primero por Entrega y luego por Paquete. AsÃƒÂ­ ves con claridad a quÃƒÂ© parte del programa le estÃƒÂ¡s cargando avance.'),
        ('Colores', 'Los colores suaves ahora distinguen Entrega 1 a 5. Ya no se usan para el estado P6.'),
        ('Control previo a carga', 'Antes de cargar, revisar que cada fila intervenida tenga sentido en Entrega, Paquete y fechas reales.'),
    ]
    ws.append(['Campo', 'Detalle'])
    style_header(ws, ['Campo', 'Detalle'])
    for item in rows:
        ws.append(list(item))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGN
    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 110


def build_meta_sheet(wb: Workbook, proj: sqlite3.Row, args, week_start: date, week_end: date, count_rows: int) -> None:
    ws = wb.create_sheet('Meta')
    ws.append(['Campo', 'Valor'])
    style_header(ws, ['Campo', 'Valor'])
    ws.append(['PROJ_ID', proj['PROJ_ID']])
    ws.append(['Programa', proj['PROJ_SHORT_NAME'] or ''])
    ws.append(['Semana ISO solicitada', args.iso_week])
    ws.append(['Desde', str(week_start)])
    ws.append(['Hasta', str(week_end)])
    ws.append(['LAST_SCHEDULE_DATE', proj['LAST_SCHEDULE_DATE'] or ''])
    ws.append(['Filas generadas', count_rows])
    ws.append(['Criterio', 'Actividades task con HH labor y solape de fechas plan/labor con la semana ISO'])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGN
    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 90


def main() -> None:
    args = parse_args()
    week_start, week_end = iso_week_bounds(args.iso_week)
    dt_start = datetime.combine(week_start, datetime.min.time())
    dt_end = datetime.combine(week_end, datetime.max.time().replace(microsecond=0))

    con = open_db(args.db)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    proj = cur.execute(
        '''
        SELECT PROJ_ID, PROJ_SHORT_NAME, LAST_SCHEDULE_DATE, DEF_COMPLETE_PCT_TYPE
        FROM PROJECT
        WHERE PROJ_ID = ?
        ''',
        (args.proj_id,),
    ).fetchone()
    if not proj:
        raise SystemExit(f'PROJ_ID no encontrado: {args.proj_id}')

    _, chain_map = build_wbs_maps(cur, args.proj_id)

    rows = cur.execute(
        '''
        SELECT
            t.TASK_ID,
            t.TASK_CODE,
            t.TASK_NAME,
            t.STATUS_CODE,
            t.WBS_ID,
            COALESCE(t.COMPLETE_PCT_TYPE, p.DEF_COMPLETE_PCT_TYPE) AS COMPLETE_PCT_TYPE,
            t.CLNDR_ID,
            t.TARGET_START_DATE,
            t.TARGET_END_DATE,
            COALESCE(SUM(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN tr.TARGET_QTY ELSE 0 END), 0) AS BAC_HH,
            COALESCE(SUM(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN COALESCE(tr.ACT_REG_QTY,0) + COALESCE(tr.ACT_OT_QTY,0) ELSE 0 END), 0) AS EV_HH,
            COALESCE(SUM(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN COALESCE(tr.REMAIN_QTY,0) ELSE 0 END), 0) AS ETC_HH,
            MIN(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN tr.TARGET_START_DATE END) AS LABOR_START,
            MAX(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN tr.TARGET_END_DATE END) AS LABOR_END,
            SUM(CASE WHEN tr.RSRC_TYPE='RT_Labor' THEN 1 ELSE 0 END) AS LABOR_ROWS
        FROM TASK t
        JOIN PROJECT p ON p.PROJ_ID = t.PROJ_ID
        LEFT JOIN TASKRSRC tr ON tr.PROJ_ID = t.PROJ_ID AND tr.TASK_ID = t.TASK_ID
        WHERE t.PROJ_ID = ?
          AND COALESCE(t.TASK_TYPE, '') NOT IN ('TT_Mile', 'TT_FinMile', 'TT_LOE', 'TT_WBS')
        GROUP BY
            t.TASK_ID, t.TASK_CODE, t.TASK_NAME, t.STATUS_CODE, t.WBS_ID,
            COALESCE(t.COMPLETE_PCT_TYPE, p.DEF_COMPLETE_PCT_TYPE),
            t.CLNDR_ID, t.TARGET_START_DATE, t.TARGET_END_DATE
        ORDER BY COALESCE(t.WBS_ID, 0), t.TASK_CODE
        ''',
        (args.proj_id,),
    ).fetchall()
    con.close()

    selected: list[dict[str, Any]] = []
    for row in rows:
        if float(row['BAC_HH'] or 0) <= 0 or int(row['LABOR_ROWS'] or 0) <= 0:
            continue
        plan_start = parse_db_dt(row['LABOR_START']) or parse_db_dt(row['TARGET_START_DATE'])
        plan_end = parse_db_dt(row['LABOR_END']) or parse_db_dt(row['TARGET_END_DATE'])
        if not overlaps(plan_start, plan_end, dt_start, dt_end):
            continue
        parts = chain_map.get(int(row['WBS_ID'])) if row['WBS_ID'] is not None else []
        entrega, paquete, ruta = derive_structure(parts or [])
        selected.append({
            'task_code': row['TASK_CODE'],
            'task_name': row['TASK_NAME'],
            'status_code': row['STATUS_CODE'] or '',
            'entrega': entrega,
            'paquete': paquete,
            'ruta': ruta,
            'pct_type': row['COMPLETE_PCT_TYPE'] or '',
            'calendar_id': row['CLNDR_ID'] or '',
            'plan_start': plan_start.strftime('%Y-%m-%d %H:%M:%S') if plan_start else '',
            'plan_end': plan_end.strftime('%Y-%m-%d %H:%M:%S') if plan_end else '',
            'bac_hh': float(row['BAC_HH'] or 0),
            'ev_hh': float(row['EV_HH'] or 0),
            'etc_hh': float(row['ETC_HH'] or 0),
        })

    selected.sort(key=lambda r: (r['entrega'], r['paquete'], r['task_code']))
    unique_deliveries = [d for d in dict.fromkeys(r['entrega'] for r in selected if r['entrega'])]
    delivery_palette = {delivery: DELIVERY_FILLS[idx % len(DELIVERY_FILLS)] for idx, delivery in enumerate(unique_deliveries)}

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet
    ws.append(HEADERS)
    style_header(ws, HEADERS)

    for item in selected:
        status_label = item['status_code'].replace('TK_', '') if item['status_code'] else ''
        ws.append([
            item['task_code'],
            item['task_name'],
            status_label,
            item['entrega'],
            item['paquete'],
            item['ruta'],
            item['pct_type'],
            item['calendar_id'],
            item['plan_start'],
            item['plan_end'],
            item['bac_hh'],
            item['ev_hh'],
            item['etc_hh'],
            '',
            '',
            '',
            '',
        ])
        row_idx = ws.max_row
        apply_delivery_fill(ws, row_idx, item['entrega'], delivery_palette)
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = WRAP_ALIGN

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{ws.max_row}'
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 85
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 22
    ws.column_dimensions['P'].width = 22
    ws.column_dimensions['Q'].width = 28
    auto_fit(ws)

    build_instructions_sheet(wb)
    build_meta_sheet(wb, proj, args, week_start, week_end, len(selected))

    wb.save(out_path)
    print(f'OUT_XLSX={out_path}')
    print(f'ROWS={len(selected)}')
    print(f'PROJ_ID={proj["PROJ_ID"]}')
    print(f'PROG={proj["PROJ_SHORT_NAME"]}')
    print(f'WEEK_START={week_start}')
    print(f'WEEK_END={week_end}')


if __name__ == '__main__':
    main()
