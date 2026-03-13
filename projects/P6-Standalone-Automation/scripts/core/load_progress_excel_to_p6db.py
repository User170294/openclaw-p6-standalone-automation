#!/usr/bin/env python3
import argparse
import csv
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any
import sys

from openpyxl import load_workbook

SCRIPTS_ROOT = Path(__file__).resolve().parents[1]
if str(SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_ROOT))

from p6_utils import now_str, open_db

DEFAULT_SHEET = 'Revision_Avance_W012'
DEFAULT_HEADERS = {
    'task_code': 'Actividad ID',
    'task_name': 'Actividad',
    'status': 'Estado P6',
    'pct_type_sheet': '% Complete Type',
    'calendar_id_sheet': 'Calendar ID',
    'bac_hh': 'BAC HH',
    'ev_hh': 'EV HH',
    'etc_hh': 'ETC HH',
    'pct_to_load': '% avance a cargar',
    'actual_start': 'Fecha inicio real a cargar',
    'actual_end_if_100': 'Fecha tÃ©rmino si 100%',
}


@dataclass
class CandidateRow:
    excel_row: int
    task_code: str
    task_name: str
    pct_to_load: float
    actual_start: datetime | None
    actual_end: datetime | None
    sheet_pct_type: str | None
    sheet_calendar_id: str | None
    sheet_status: str | None
    issues: list[str]


@dataclass
class TaskContext:
    task_id: int
    task_code: str
    task_name: str
    proj_id: int
    status_code: str | None
    complete_pct_type: str | None
    proj_complete_pct_type: str | None
    clndr_id: Any
    target_start_date: datetime | None
    target_end_date: datetime | None
    act_start_date: datetime | None
    act_end_date: datetime | None


def parse_args():
    p = argparse.ArgumentParser(description='Carga segura avances Excel -> DB SQLite Primavera P6.')
    p.add_argument('--db', required=True, help='Ruta a PPMDBSQLite .db')
    p.add_argument('--xlsx', required=True, help='Excel de revisiÃ³n/carga')
    p.add_argument('--proj-id', type=int, required=True, help='PROJ_ID objetivo')
    p.add_argument('--sheet', default=DEFAULT_SHEET, help='Nombre de hoja Excel')
    p.add_argument('--out-dir', default='projects/P6-Standalone-Automation/data', help='Carpeta para reportes')
    p.add_argument('--apply', action='store_true', help='Aplica cambios. Sin esto, solo dry-run.')
    p.add_argument('--backup-dir', default='', help='Carpeta opcional para respaldos .db')
    p.add_argument('--default-day-start', default='08:00', help='Hora inicio por defecto para fechas sin hora')
    p.add_argument('--default-day-end', default='18:00', help='Hora tÃ©rmino por defecto para fechas sin hora')
    p.add_argument('--task-code-col', default=DEFAULT_HEADERS['task_code'])
    p.add_argument('--task-name-col', default=DEFAULT_HEADERS['task_name'])
    p.add_argument('--status-col', default=DEFAULT_HEADERS['status'])
    p.add_argument('--pct-type-col', default=DEFAULT_HEADERS['pct_type_sheet'])
    p.add_argument('--calendar-id-col', default=DEFAULT_HEADERS['calendar_id_sheet'])
    p.add_argument('--pct-col', default=DEFAULT_HEADERS['pct_to_load'])
    p.add_argument('--actual-start-col', default=DEFAULT_HEADERS['actual_start'])
    p.add_argument('--actual-end-col', default=DEFAULT_HEADERS['actual_end_if_100'])
    return p.parse_args()


def parse_clock(value: str) -> time:
    return datetime.strptime(value, '%H:%M').time()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == '')


def clean_str(value: Any) -> str:
    return '' if value is None else str(value).strip()


def parse_pct(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        num = float(value)
    else:
        txt = str(value).strip().replace('%', '').replace(',', '.')
        num = float(txt)
    if 0 <= num <= 1:
        num *= 100.0
    return round(num, 6)


def parse_excel_datetime(value: Any) -> datetime | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return datetime(value.year, value.month, value.day)
    txt = str(value).strip()
    for fmt in (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d-%m-%Y %H:%M:%S',
        '%d-%m-%Y %H:%M',
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M',
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%d/%m/%Y',
    ):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            pass
    raise ValueError(f'Fecha no reconocida: {value!r}')


def parse_db_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    return datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S')


def format_db_datetime(value: datetime | None) -> str | None:
    return None if value is None else value.strftime('%Y-%m-%d %H:%M:%S')


def has_explicit_time(value: datetime | None) -> bool:
    return bool(value and (value.hour or value.minute or value.second or value.microsecond))


def combine_date_and_time(day: datetime, clock: time) -> datetime:
    return datetime(day.year, day.month, day.day, clock.hour, clock.minute, clock.second)


def normalize_same_day(actual_start: datetime | None, actual_end: datetime | None, target_start: datetime | None, target_end: datetime | None, default_start: time, default_end: time) -> tuple[datetime | None, datetime | None, list[str]]:
    notes: list[str] = []
    if actual_start is None and actual_end is None:
        return None, None, notes

    if actual_start and actual_end and actual_start.date() == actual_end.date() and not has_explicit_time(actual_start) and not has_explicit_time(actual_end):
        if target_start and target_end and target_start.date() == target_end.date():
            actual_start = combine_date_and_time(actual_start, target_start.time())
            actual_end = combine_date_and_time(actual_end, target_end.time())
            notes.append('same_day_dates_normalized_from_target_times')
        else:
            actual_start = combine_date_and_time(actual_start, default_start)
            actual_end = combine_date_and_time(actual_end, default_end)
            notes.append('same_day_dates_normalized_from_default_times')
    elif actual_start and not has_explicit_time(actual_start):
        if target_start:
            actual_start = combine_date_and_time(actual_start, target_start.time())
            notes.append('actual_start_time_inherited_from_target_start')
        else:
            actual_start = combine_date_and_time(actual_start, default_start)
            notes.append('actual_start_time_defaulted')
    elif actual_end and not has_explicit_time(actual_end):
        if target_end:
            actual_end = combine_date_and_time(actual_end, target_end.time())
            notes.append('actual_end_time_inherited_from_target_end')
        else:
            actual_end = combine_date_and_time(actual_end, default_end)
            notes.append('actual_end_time_defaulted')
    return actual_start, actual_end, notes


def load_candidates(args) -> tuple[list[CandidateRow], list[str]]:
    wb = load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f'Hoja no encontrada: {args.sheet}')
    ws = wb[args.sheet]
    headers = [c.value for c in ws[1]]
    idx = {str(h).strip(): pos for pos, h in enumerate(headers) if h is not None}

    required = [
        args.task_code_col,
        args.task_name_col,
        args.status_col,
        args.pct_col,
        args.actual_start_col,
        args.actual_end_col,
        args.pct_type_col,
        args.calendar_id_col,
    ]
    missing = [h for h in required if h not in idx]
    if missing:
        raise SystemExit(f'Faltan columnas requeridas en Excel: {missing}')

    parsed: list[CandidateRow] = []
    workbook_issues: list[str] = []

    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        task_code = clean_str(row[idx[args.task_code_col]])
        if not task_code:
            continue
        task_name = clean_str(row[idx[args.task_name_col]])
        pct = parse_pct(row[idx[args.pct_col]])
        actual_start = parse_excel_datetime(row[idx[args.actual_start_col]]) if not is_blank(row[idx[args.actual_start_col]]) else None
        actual_end = parse_excel_datetime(row[idx[args.actual_end_col]]) if not is_blank(row[idx[args.actual_end_col]]) else None
        sheet_pct_type = clean_str(row[idx[args.pct_type_col]]) or None
        sheet_calendar_id = clean_str(row[idx[args.calendar_id_col]]) or None
        sheet_status = clean_str(row[idx[args.status_col]]) or None

        if pct is None and actual_start is None and actual_end is None:
            continue

        issues: list[str] = []
        if pct is None:
            issues.append('missing_pct_to_load')
        elif pct < 0 or pct > 100:
            issues.append('pct_out_of_range')

        if pct == 100 and actual_end is None:
            issues.append('missing_actual_end_for_100pct')
        if pct is not None and 0 < pct < 100 and actual_end is not None:
            issues.append('actual_end_present_for_partial_progress')

        parsed.append(CandidateRow(
            excel_row=excel_row,
            task_code=task_code,
            task_name=task_name,
            pct_to_load=pct or 0.0,
            actual_start=actual_start,
            actual_end=actual_end,
            sheet_pct_type=sheet_pct_type,
            sheet_calendar_id=sheet_calendar_id,
            sheet_status=sheet_status,
            issues=issues,
        ))

    if not parsed:
        workbook_issues.append('no_candidate_rows_detected')
    return parsed, workbook_issues


def fetch_task_context(cur: sqlite3.Cursor, proj_id: int, task_code: str) -> TaskContext | None:
    row = cur.execute(
        '''
        SELECT
            t.TASK_ID,
            t.TASK_CODE,
            t.TASK_NAME,
            t.PROJ_ID,
            t.STATUS_CODE,
            t.COMPLETE_PCT_TYPE,
            p.DEF_COMPLETE_PCT_TYPE AS PROJ_COMPLETE_PCT_TYPE,
            t.CLNDR_ID,
            t.TARGET_START_DATE,
            t.TARGET_END_DATE,
            t.ACT_START_DATE,
            t.ACT_END_DATE
        FROM TASK t
        JOIN PROJECT p ON p.PROJ_ID = t.PROJ_ID
        WHERE t.PROJ_ID = ? AND t.TASK_CODE = ?
        ''',
        (proj_id, task_code),
    ).fetchone()
    if not row:
        return None
    return TaskContext(
        task_id=row[0],
        task_code=row[1],
        task_name=row[2],
        proj_id=row[3],
        status_code=row[4],
        complete_pct_type=row[5],
        proj_complete_pct_type=row[6],
        clndr_id=row[7],
        target_start_date=parse_db_datetime(row[8]),
        target_end_date=parse_db_datetime(row[9]),
        act_start_date=parse_db_datetime(row[10]),
        act_end_date=parse_db_datetime(row[11]),
    )


def backup_db(db_path: Path, backup_dir: Path | None) -> Path:
    stamp = now_str().replace('-', '').replace(':', '').replace(' ', '_')
    if backup_dir is None:
        backup_path = db_path.with_suffix(db_path.suffix + f'.BACKUP_{stamp}')
    else:
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f'{db_path.name}.BACKUP_{stamp}'
    shutil.copy2(db_path, backup_path)
    return backup_path


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main():
    args = parse_args()
    db_path = Path(args.db)
    xlsx_path = Path(args.xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    candidates, workbook_issues = load_candidates(args)

    con = open_db(db_path)
    cur = con.cursor()

    default_start = parse_clock(args.default_day_start)
    default_end = parse_clock(args.default_day_end)
    stamp = now_str().replace('-', '').replace(':', '').replace(' ', '_')

    preview_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []
    applied_count = 0
    backup_path: Path | None = None

    try:
        if args.apply:
            backup_path = backup_db(db_path, Path(args.backup_dir) if args.backup_dir else None)
            cur.execute('BEGIN')

        for cand in candidates:
            ctx = fetch_task_context(cur, args.proj_id, cand.task_code)
            issues = list(cand.issues)
            notes: list[str] = []

            if ctx is None:
                issues.append('task_not_found_in_db')
                error_rows.append({
                    'excel_row': cand.excel_row,
                    'task_code': cand.task_code,
                    'task_name': cand.task_name,
                    'issues': ';'.join(issues),
                })
                continue

            if cand.sheet_calendar_id and str(ctx.clndr_id) != cand.sheet_calendar_id:
                issues.append(f'calendar_id_mismatch_sheet={cand.sheet_calendar_id}_db={ctx.clndr_id}')

            effective_pct_type = ctx.complete_pct_type or ctx.proj_complete_pct_type
            if ctx.complete_pct_type and ctx.proj_complete_pct_type and ctx.complete_pct_type != ctx.proj_complete_pct_type:
                notes.append(f'task_pct_type_differs_from_project:{ctx.complete_pct_type}->{ctx.proj_complete_pct_type}')
            if effective_pct_type != (ctx.proj_complete_pct_type or effective_pct_type):
                notes.append('task_pct_type_will_align_to_project')
                effective_pct_type = ctx.proj_complete_pct_type or effective_pct_type

            actual_start = cand.actual_start
            actual_end = cand.actual_end
            actual_start, actual_end, norm_notes = normalize_same_day(actual_start, actual_end, ctx.target_start_date, ctx.target_end_date, default_start, default_end)
            notes.extend(norm_notes)

            if actual_start and actual_end and actual_end < actual_start:
                issues.append('actual_end_before_actual_start')
            if cand.pct_to_load > 0 and actual_start is None and ctx.act_start_date is None:
                issues.append('actual_start_missing_after_normalization')
            if cand.pct_to_load == 100 and actual_end is None:
                issues.append('actual_end_missing_after_normalization')
            if cand.pct_to_load > 0 and actual_start is None and ctx.act_start_date is not None:
                notes.append('actual_start_reused_from_db')

            tr = cur.execute(
                '''
                SELECT
                    COUNT(*) AS rows_n,
                    COALESCE(SUM(CASE WHEN RSRC_TYPE='RT_Labor' THEN TARGET_QTY ELSE 0 END), 0) AS target_qty,
                    COALESCE(SUM(CASE WHEN RSRC_TYPE='RT_Labor' THEN TARGET_COST ELSE 0 END), 0) AS target_cost
                FROM TASKRSRC
                WHERE PROJ_ID = ? AND TASK_ID = ?
                ''',
                (args.proj_id, ctx.task_id),
            ).fetchone()
            target_qty = float(tr['target_qty'])
            target_cost = float(tr['target_cost'])
            if tr['rows_n'] == 0 or target_qty <= 0:
                issues.append('no_labor_taskrsrc_rows')

            act_qty = round(target_qty * (cand.pct_to_load / 100.0), 6)
            remain_qty = round(max(0.0, target_qty - act_qty), 6)
            act_cost = round(target_cost * (cand.pct_to_load / 100.0), 6)
            remain_cost = round(max(0.0, target_cost - act_cost), 6)

            new_status = 'TK_Complete' if cand.pct_to_load >= 100 else 'TK_Active'
            task_start = ctx.act_start_date or actual_start or ctx.target_start_date
            task_end = actual_end if cand.pct_to_load >= 100 else None

            if issues:
                error_rows.append({
                    'excel_row': cand.excel_row,
                    'task_code': cand.task_code,
                    'task_name': cand.task_name,
                    'issues': ';'.join(issues),
                })
                continue

            preview_rows.append({
                'excel_row': cand.excel_row,
                'task_code': cand.task_code,
                'task_name': ctx.task_name,
                'pct_to_load': f'{cand.pct_to_load:.4f}',
                'status_before': ctx.status_code or '',
                'status_after': new_status,
                'pct_type_before': ctx.complete_pct_type or '',
                'pct_type_after': effective_pct_type or '',
                'calendar_id': ctx.clndr_id,
                'actual_start_after': format_db_datetime(task_start) or '',
                'actual_end_after': format_db_datetime(task_end) or '',
                'target_qty': f'{target_qty:.4f}',
                'act_qty_after': f'{act_qty:.4f}',
                'remain_qty_after': f'{remain_qty:.4f}',
                'target_cost': f'{target_cost:.4f}',
                'act_cost_after': f'{act_cost:.4f}',
                'remain_cost_after': f'{remain_cost:.4f}',
                'notes': ';'.join(notes),
            })

            if not args.apply:
                continue

            cur.execute(
                '''
                UPDATE TASKRSRC
                SET ACT_REG_QTY = CASE WHEN RSRC_TYPE='RT_Labor' THEN ROUND(TARGET_QTY * ?, 6) ELSE ACT_REG_QTY END,
                    REMAIN_QTY = CASE WHEN RSRC_TYPE='RT_Labor' THEN ROUND(MAX(0, TARGET_QTY - (TARGET_QTY * ?)), 6) ELSE REMAIN_QTY END,
                    ACT_REG_COST = CASE WHEN RSRC_TYPE='RT_Labor' THEN ROUND(TARGET_COST * ?, 6) ELSE ACT_REG_COST END,
                    REMAIN_COST = CASE WHEN RSRC_TYPE='RT_Labor' THEN ROUND(MAX(0, TARGET_COST - (TARGET_COST * ?)), 6) ELSE REMAIN_COST END,
                    ACT_START_DATE = ?,
                    ACT_END_DATE = ?,
                    UPDATE_DATE = ?,
                    UPDATE_USER = 'openclaw'
                WHERE PROJ_ID = ? AND TASK_ID = ?
                ''',
                (
                    cand.pct_to_load / 100.0,
                    cand.pct_to_load / 100.0,
                    cand.pct_to_load / 100.0,
                    cand.pct_to_load / 100.0,
                    format_db_datetime(task_start),
                    format_db_datetime(task_end),
                    now_str(),
                    args.proj_id,
                    ctx.task_id,
                ),
            )

            cur.execute(
                '''
                UPDATE TASK
                SET STATUS_CODE = ?,
                    COMPLETE_PCT_TYPE = ?,
                    ACT_START_DATE = ?,
                    ACT_END_DATE = ?,
                    TARGET_WORK_QTY = COALESCE((SELECT SUM(tr.TARGET_QTY) FROM TASKRSRC tr WHERE tr.PROJ_ID=TASK.PROJ_ID AND tr.TASK_ID=TASK.TASK_ID AND tr.RSRC_TYPE='RT_Labor'), 0),
                    ACT_WORK_QTY = COALESCE((SELECT SUM(tr.ACT_REG_QTY) FROM TASKRSRC tr WHERE tr.PROJ_ID=TASK.PROJ_ID AND tr.TASK_ID=TASK.TASK_ID AND tr.RSRC_TYPE='RT_Labor'), 0),
                    REMAIN_WORK_QTY = COALESCE((SELECT SUM(tr.REMAIN_QTY) FROM TASKRSRC tr WHERE tr.PROJ_ID=TASK.PROJ_ID AND tr.TASK_ID=TASK.TASK_ID AND tr.RSRC_TYPE='RT_Labor'), 0),
                    UPDATE_DATE = ?,
                    UPDATE_USER = 'openclaw'
                WHERE PROJ_ID = ? AND TASK_ID = ?
                ''',
                (
                    new_status,
                    effective_pct_type,
                    format_db_datetime(task_start),
                    format_db_datetime(task_end),
                    now_str(),
                    args.proj_id,
                    ctx.task_id,
                ),
            )
            applied_count += 1

        if args.apply:
            con.commit()
    except Exception:
        if args.apply:
            con.rollback()
        raise
    finally:
        con.close()

    preview_csv = out_dir / f'load_progress_preview_{args.proj_id}_{stamp}.csv'
    errors_csv = out_dir / f'load_progress_errors_{args.proj_id}_{stamp}.csv'
    summary_md = out_dir / f'load_progress_summary_{args.proj_id}_{stamp}.md'

    write_csv(preview_csv, preview_rows, [
        'excel_row', 'task_code', 'task_name', 'pct_to_load', 'status_before', 'status_after', 'pct_type_before',
        'pct_type_after', 'calendar_id', 'actual_start_after', 'actual_end_after', 'target_qty', 'act_qty_after',
        'remain_qty_after', 'target_cost', 'act_cost_after', 'remain_cost_after', 'notes'
    ])
    write_csv(errors_csv, error_rows, ['excel_row', 'task_code', 'task_name', 'issues'])

    with summary_md.open('w', encoding='utf-8') as f:
        f.write(f'# Safe Load Summary ({stamp})\n\n')
        f.write(f'- DB: `{db_path}`\n')
        f.write(f'- XLSX: `{xlsx_path}`\n')
        f.write(f'- PROJ_ID: `{args.proj_id}`\n')
        f.write(f'- Sheet: `{args.sheet}`\n')
        f.write(f'- Apply: `{args.apply}`\n')
        f.write(f'- Workbook issues: `{workbook_issues}`\n')
        f.write(f'- Candidate rows: **{len(candidates)}**\n')
        f.write(f'- Ready/preview rows: **{len(preview_rows)}**\n')
        f.write(f'- Error rows: **{len(error_rows)}**\n')
        f.write(f'- Applied rows: **{applied_count}**\n')
        if backup_path:
            f.write(f'- Backup DB: `{backup_path}`\n')
        f.write(f'\n## Archivos\n')
        f.write(f'- Preview CSV: `{preview_csv}`\n')
        f.write(f'- Errors CSV: `{errors_csv}`\n')

    print(f'CANDIDATES={len(candidates)}')
    print(f'PREVIEW_ROWS={len(preview_rows)}')
    print(f'ERROR_ROWS={len(error_rows)}')
    print(f'APPLIED_ROWS={applied_count}')
    if backup_path:
        print(f'BACKUP_DB={backup_path}')
    print(f'PREVIEW_CSV={preview_csv}')
    print(f'ERRORS_CSV={errors_csv}')
    print(f'SUMMARY_MD={summary_md}')


if __name__ == '__main__':
    main()