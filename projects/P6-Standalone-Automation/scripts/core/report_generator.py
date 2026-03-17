from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill

FIXED_COLUMNS = ['week', 'pv_week', 'pv_cum', 'pv_pct', 'ev_cum', 'ev_pct', 'ac_cum', 'sv', 'spi', 'forecast_cum', 'forecast_pct']


def _safe_num(value: Any) -> float | None:
    if value is None or value == '':
        return None
    try:
        return float(value)
    except Exception:
        return None


def enrich_report(report: dict[str, Any]) -> dict[str, Any]:
    rows = report.get('rows', [])
    last = rows[-1] if rows else {}
    bac = _safe_num(report.get('bac'))
    if bac is None:
        bac = sum(float(row.get('pv_week') or 0.0) for row in rows)
    pv = _safe_num(last.get('pv_cum')) or 0.0
    ev = _safe_num(last.get('ev_cum')) or 0.0
    ac = _safe_num(last.get('ac_cum')) or _safe_num(report.get('ac')) or 0.0
    sv = _safe_num(last.get('sv'))
    spi = _safe_num(last.get('spi'))
    cpi = _safe_num(report.get('cpi'))
    eac = _safe_num(report.get('eac'))
    etc = _safe_num(report.get('etc'))
    cv = _safe_num(report.get('cv'))
    enriched = dict(report)
    enriched['bac'] = round(bac, 4)
    enriched['kpis'] = {
        'bac': round(bac, 4),
        'pv': round(pv, 4),
        'ev': round(ev, 4),
        'ac': round(ac, 4),
        'sv': round(sv, 4) if sv is not None else None,
        'cv': round(cv, 4) if cv is not None else None,
        'spi': round(spi, 6) if spi is not None else None,
        'cpi': round(cpi, 6) if cpi is not None else None,
        'eac': round(eac, 4) if eac is not None else None,
        'etc': round(etc, 4) if etc is not None else None,
    }
    return enriched


def _fmt_num(value: Any, decimals: int = 4) -> str:
    if value is None or value == '':
        return 'N/A'
    try:
        return f'{float(value):,.{decimals}f}'
    except Exception:
        return str(value)


def _report_stem(report: dict[str, Any], meta: dict[str, Any] | None) -> str:
    mode = report.get('mode', 'logic')
    proj = (meta or {}).get('project_name') or (meta or {}).get('proj_id') or 'report'
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    proj_txt = str(proj).replace(' ', '_')
    return f'report_{proj_txt}_{mode}_{stamp}'


def render_md(report: dict[str, Any], out_path: Path, meta: dict[str, Any] | None = None) -> Path:
    report = enrich_report(report)
    kpis = report['kpis']
    title = (meta or {}).get('project_name') or 'P6 Standalone Report'
    lines = [
        f'# {title}',
        '',
        '## Estado del reporte',
        f"- Modo: `{report.get('mode', '')}`",
        f"- Nota: {report.get('note', '')}",
        '',
        '## KPIs principales',
        f"- BAC: **{_fmt_num(kpis['bac'])}**",
        f"- PV: **{_fmt_num(kpis['pv'])}**",
        f"- EV: **{_fmt_num(kpis['ev'])}**",
        f"- AC: **{_fmt_num(kpis['ac'])}**",
        f"- SV: **{_fmt_num(kpis['sv'])}**",
        f"- CV: **{_fmt_num(kpis['cv'])}**",
        f"- SPI: **{_fmt_num(kpis['spi'], 6)}**",
        f"- CPI: **{_fmt_num(kpis['cpi'], 6)}**",
        f"- EAC: **{_fmt_num(kpis['eac'])}**",
        f"- ETC: **{_fmt_num(kpis['etc'])}**",
        '',
        '## Tabla semanal',
        '| ' + ' | '.join(FIXED_COLUMNS) + ' |',
        '|' + '|'.join(['---'] * len(FIXED_COLUMNS)) + '|',
    ]
    for row in report.get('rows', []):
        vals = [str(row.get(col, '')) for col in FIXED_COLUMNS]
        lines.append('| ' + ' | '.join(vals) + ' |')
    out_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')
    return out_path


def render_xlsx(report: dict[str, Any], out_path: Path, meta: dict[str, Any] | None = None) -> Path:
    report = enrich_report(report)
    wb = Workbook()
    ws = wb.active
    ws.title = 'weekly_data'
    ws.append(FIXED_COLUMNS)
    header_fill = PatternFill(fill_type='solid', fgColor='0D0F12')
    header_font = Font(color='FFFFFF', bold=True)
    for idx, _col in enumerate(FIXED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[chr(64 + idx)].width = 16
    for row in report.get('rows', []):
        ws.append([row.get(col) for col in FIXED_COLUMNS])

    summary = wb.create_sheet('summary_chart')
    summary['A1'] = (meta or {}).get('project_name') or 'P6 Standalone Report'
    summary['A3'] = 'BAC';  summary['B3'] = report['kpis']['bac']
    summary['A4'] = 'PV';   summary['B4'] = report['kpis']['pv']
    summary['A5'] = 'EV';   summary['B5'] = report['kpis']['ev']
    summary['A6'] = 'AC';   summary['B6'] = report['kpis']['ac']
    summary['A7'] = 'SV';   summary['B7'] = report['kpis']['sv']
    summary['A8'] = 'CV';   summary['B8'] = report['kpis']['cv']
    summary['A9'] = 'SPI';  summary['B9'] = report['kpis']['spi']
    summary['A10'] = 'CPI'; summary['B10'] = report['kpis']['cpi']
    summary['A11'] = 'EAC'; summary['B11'] = report['kpis']['eac']
    summary['A12'] = 'ETC'; summary['B12'] = report['kpis']['etc']

    chart = LineChart()
    chart.title = 'Curva S'
    chart.y_axis.title = 'HH'
    chart.x_axis.title = 'Week'
    data = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=max(2, ws.max_row))
    cats = Reference(ws, min_col=1, min_row=2, max_row=max(2, ws.max_row))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    summary.add_chart(chart, 'D3')

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def generate_report(report: dict[str, Any], out_dir: str | Path, fmt: str, meta: dict[str, Any] | None = None) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f'{_report_stem(report, meta)}.{fmt}'
    if fmt == 'md':
        return render_md(report, out_path, meta)
    if fmt == 'xlsx':
        return render_xlsx(report, out_path, meta)
    raise ValueError(f'Formato de reporte no soportado: {fmt}')


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(description='Renderiza reportes finales desde el dict/JSON de pv_engine.compare().')
    ap.add_argument('--input-json', required=True)
    ap.add_argument('--out-dir', required=True)
    ap.add_argument('--format', choices=['md', 'xlsx'], required=True)
    ap.add_argument('--project-name', default='P6 Standalone Report')
    args = ap.parse_args()

    report = json.loads(Path(args.input_json).read_text(encoding='utf-8'))
    out = generate_report(report, args.out_dir, args.format, {'project_name': args.project_name})
    print(f'OUT={out}')


if __name__ == '__main__':
    main()
