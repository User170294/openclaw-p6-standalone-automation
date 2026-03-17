#!/usr/bin/env python3
import argparse
import sqlite3
from pathlib import Path
import sys
from openpyxl import load_workbook

SCRIPTS_ROOT = Path(__file__).resolve().parents[1]
if str(SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_ROOT))

from p6_utils import open_db


def parse_args():
    ap = argparse.ArgumentParser(description='Prepara workbook tecnico para carga segura desde planilla de captura.')
    ap.add_argument('--db', required=True)
    ap.add_argument('--source-xlsx', required=True)
    ap.add_argument('--out-xlsx', required=True)
    ap.add_argument('--proj-id', type=int, required=True)
    ap.add_argument('--sheet', default='Revision_Avance_W012')
    return ap.parse_args()


def main():
    args = parse_args()
    wb = load_workbook(args.source_xlsx)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f'Hoja no encontrada: {args.sheet}')
    ws = wb[args.sheet]

    headers = [c.value for c in ws[1]]
    idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None}

    insert_after = idx['Estado P6']
    ws.insert_cols(insert_after + 1, amount=2)
    ws.cell(1, insert_after + 1).value = '% Complete Type'
    ws.cell(1, insert_after + 2).value = 'Calendar ID'

    con = open_db(args.db)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    task_code_col = idx['Actividad ID']
    for r in range(2, ws.max_row + 1):
        task_code = ws.cell(r, task_code_col).value
        if not task_code:
            continue
        row = cur.execute(
            '''
            SELECT COALESCE(t.COMPLETE_PCT_TYPE, p.DEF_COMPLETE_PCT_TYPE) AS COMPLETE_PCT_TYPE,
                   t.CLNDR_ID
            FROM TASK t
            JOIN PROJECT p ON p.PROJ_ID = t.PROJ_ID
            WHERE t.PROJ_ID = ? AND t.TASK_CODE = ?
            ''',
            (args.proj_id, str(task_code).strip()),
        ).fetchone()
        if row:
            ws.cell(r, insert_after + 1).value = row['COMPLETE_PCT_TYPE']
            ws.cell(r, insert_after + 2).value = row['CLNDR_ID']

    con.close()
    out_path = Path(args.out_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'OUT_XLSX={out_path}')


if __name__ == '__main__':
    main()
