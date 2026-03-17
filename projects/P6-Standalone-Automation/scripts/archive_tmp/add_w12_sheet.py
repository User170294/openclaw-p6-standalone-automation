"""
Agrega hoja W12 - Plan de Trabajo al Excel EVM existente
Agrupado por TAG (unidad) y subsistema
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r'C:\Users\josej\Desktop\EVM_OT1844_W11.xlsx'

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)
def brd():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HDR_FILL  = fill("1F3864"); HDR_FONT  = font(bold=True, color="FFFFFF", size=10)
TAG_FILL  = fill("2E4057"); TAG_FONT  = font(bold=True, color="FFFFFF", size=10)
SUB_FILL  = fill("4472C4"); SUB_FONT  = font(bold=True, color="FFFFFF", size=10)
INFO_FILL = fill("CFE2F3")
WARN_FILL = fill("FCE5CD")
OK_FILL   = fill("D9EAD3")
RED_FILL  = fill("F4CCCC")
RED_FONT  = font(bold=True, color="CC0000", size=10)
GRN_FONT  = font(bold=True, color="006600", size=10)
NRM_FONT  = font(size=10)
BLD_FONT  = font(bold=True, size=10)

def c(ws, row, col, val, fnt=None, flll=None, aln=CTR, border=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = fnt or NRM_FONT
    cell.alignment = aln
    if flll: cell.fill = flll
    if border: cell.border = brd()
    return cell

# Datos por TAG / subsistema
TAGS = {
    "TAG 2225-ZM-01": {
        "hh_rem": 30.6,
        "subsistemas": {
            "Parrilla de Piso (FRP)": [
                ("A2830","Trazado y Corte de FRP",              54, 32.4, 21.6, 60.0, "EN CURSO", "17-Mar"),
                ("A2840","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","18-Mar"),
            ]
        }
    },
    "TAG 2225-ZM-02": {
        "hh_rem": 30.6,
        "subsistemas": {
            "Parrilla de Piso (FRP)": [
                ("A3150","Trazado y Corte de FRP",              54, 32.4, 21.6, 60.0, "EN CURSO", "17-Mar"),
                ("A3160","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","17-Mar"),
            ]
        }
    },
    "TAG 2225-ZM-03": {
        "hh_rem": 774.9,
        "subsistemas": {
            "Bastidor": [
                ("A3280","Corte y Fab. Vigas y Perfil Z",      162,135.0, 27.0, 83.3, "EN CURSO", "16-Mar"),
                ("A3290","Corte y Fab. Consolas Exteriores",    36, 18.0, 18.0, 50.0, "EN CURSO", "16-Mar"),
                ("A3300","Corte y Fab. Consolas Centrales",     36, 18.0, 18.0, 50.0, "EN CURSO", "16-Mar"),
                ("A3320","Armado y Remates Soldadura Bastidor", 54,  0.0, 54.0,  0.0, "NO INICIO","17-Mar"),
                ("A3330","Inspección Soldadura (VT/LP)",         9,  0.0,  9.0,  0.0, "NO INICIO","17-Mar"),
                ("A3340","Granallado / Prep. Superficie",       18,  0.0, 18.0,  0.0, "NO INICIO","18-Mar"),
                ("A3350","Aplic. Imprimante/Enlace/Terminación",54,  0.0, 54.0,  0.0, "NO INICIO","21-Mar"),
                ("A3360","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","21-Mar"),
            ],
            "Bandeja Perimetral SS316L": [
                ("A3390","Armado, Soldadura y Rect. Bandejas", 108,  0.0,108.0,  0.0, "NO INICIO","18-Mar"),
                ("A3400","Terminaciones Superficiales SS316L",  54,  0.0, 54.0,  0.0, "NO INICIO","20-Mar"),
                ("A3410","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","20-Mar"),
            ],
            "Bandeja Central SS316L": [
                ("A3530","Armado, Soldadura y Rect. Bandejas", 108,  0.0,108.0,  0.0, "NO INICIO","18-Mar"),
                ("A3540","Terminaciones Superficiales SS316L",  54,  0.0, 54.0,  0.0, "NO INICIO","20-Mar"),
                ("A3550","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","20-Mar"),
            ],
            "Barandas": [
                ("A3420","Fab. Baranda Radial Removible",      189, 94.5, 94.5, 50.0, "EN CURSO", "19-Mar"),
                ("A3430","Aplic. Imprimante/Enlace/Terminación",54,  0.0, 54.0,  0.0, "NO INICIO","22-Mar"),
                ("A3440","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","22-Mar"),
            ],
            "Manifold Distribuidor": [
                ("A3470","Fab. Cuerpo Central (Spool/Cabezal)", 36, 28.8,  7.2, 80.0, "EN CURSO", "16-Mar"),
                ("A3480","Fab. Brazos Radiales",                18,  0.0, 18.0,  0.0, "NO INICIO","17-Mar"),
                ("A3490","Fab. Conjunto Válvula/Boquilla",      18,  0.0, 18.0,  0.0, "NO INICIO","18-Mar"),
                ("A3500","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","19-Mar"),
            ],
            "Parrilla FRP": [
                ("A3450","Trazado y Corte de FRP",              54, 37.8, 16.2, 70.0, "EN CURSO", "20-Mar"),
            ],
        }
    },
    "TAG 2225-ZM-04": {
        "hh_rem": 672.3,
        "subsistemas": {
            "Bastidor": [
                ("A3580","Corte y Fab. Vigas y Perfil Z",      162, 81.0, 81.0, 50.0, "EN CURSO", "18-Mar"),
                ("A3590","Corte y Fab. Consolas Exteriores",    36, 10.8, 25.2, 30.0, "EN CURSO", "17-Mar"),
                ("A3600","Corte y Fab. Consolas Centrales",     36, 10.8, 25.2, 30.0, "EN CURSO", "18-Mar"),
                ("A3620","Armado y Remates Soldadura Bastidor", 54,  0.0, 54.0,  0.0, "NO INICIO","20-Mar"),
            ],
            "Bandeja Perimetral SS316L": [
                ("A3690","Armado, Soldadura y Rect. Bandejas", 108,  0.0,108.0,  0.0, "NO INICIO","18-Mar"),
                ("A3700","Terminaciones Superficiales SS316L",  54,  0.0, 54.0,  0.0, "NO INICIO","20-Mar"),
            ],
            "Bandeja Central SS316L": [
                ("A3830","Armado, Soldadura y Rect. Bandejas", 108,  0.0,108.0,  0.0, "NO INICIO","18-Mar"),
                ("A3840","Terminaciones Superficiales SS316L",  54,  0.0, 54.0,  0.0, "NO INICIO","20-Mar"),
            ],
            "Barandas": [
                ("A3720","Fab. Baranda Radial Removible",      189, 94.5, 94.5, 50.0, "EN CURSO", "19-Mar"),
            ],
            "Manifold Distribuidor": [
                ("A3770","Fab. Cuerpo Central (Spool/Cabezal)", 36, 28.8,  7.2, 80.0, "EN CURSO", "16-Mar"),
                ("A3780","Fab. Brazos Radiales",                18,  0.0, 18.0,  0.0, "NO INICIO","17-Mar"),
                ("A3790","Fab. Conjunto Válvula/Boquilla",      18,  0.0, 18.0,  0.0, "NO INICIO","18-Mar"),
                ("A3800","Control de Calidad (QC)",              9,  0.0,  9.0,  0.0, "NO INICIO","19-Mar"),
            ],
            "Parrilla FRP": [
                ("A3750","Trazado y Corte de FRP",              54, 37.8, 16.2, 70.0, "EN CURSO", "20-Mar"),
            ],
        }
    },
    "TAG 2225-ZM-05": {
        "hh_rem": 36.0,
        "subsistemas": {
            "Bastidor": [
                ("A3890","Corte y Fab. Consolas Exteriores",    36,  0.0, 36.0,  0.0, "NO INICIO","20-Mar"),
            ]
        }
    },
}

def add_w12(wb):
    if "W12 - Plan de Trabajo" in wb.sheetnames:
        del wb["W12 - Plan de Trabajo"]
    ws = wb.create_sheet("W12 - Plan de Trabajo")
    ws.freeze_panes = "A5"

    # Título
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "PLAN DE TRABAJO W12 — OT-1844 | 16 al 22 Mar 2026 | Actividades a completar para cumplir Forecast"
    cell.font = font(bold=True, size=12, color="1F3864")
    cell.alignment = CTR; cell.fill = INFO_FILL
    ws.row_dimensions[1].height = 22

    # Resumen KPI
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = ("  Total: 41 actividades pendientes  |  1,544.4 HH remaining  |  "
                  "~309 HH/día requeridas (5 días hábiles)  |  "
                  "Forecast W12: 6,541.2 HH acum → 80.76%")
    cell.font = BLD_FONT; cell.alignment = LFT; cell.fill = INFO_FILL
    ws.row_dimensions[2].height = 16

    # Alerta
    ws.merge_cells("A3:H3")
    cell = ws["A3"]
    cell.value = ("  ⚠  CRITICO: 4 actividades de Armado/Soldadura Bandejas SS316L (432 HH) "
                  "deben iniciar el 16-Mar. Las 2 Barandas Radiales (189 HH c/u) deben cerrar antes del 19-Mar.")
    cell.font = RED_FONT; cell.alignment = LFT; cell.fill = WARN_FILL
    ws.row_dimensions[3].height = 18

    # Headers
    hdrs = ["TAG","Subsistema","Código","Actividad","HH Target","HH Act.","HH Rem.","Avance %","Estado","Early End"]
    wids = [16, 22, 9, 45, 10, 10, 10, 10, 11, 10]
    for i,(h,w) in enumerate(zip(hdrs, wids), 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill=HDR_FILL; cell.font=HDR_FONT; cell.alignment=CTR; cell.border=brd()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28

    row = 5
    for tag_name, tag_data in TAGS.items():
        # Fila TAG
        ws.merge_cells(f"A{row}:J{row}")
        cell = ws[f"A{row}"]
        cell.value = f"  {tag_name}  —  {tag_data['hh_rem']:.1f} HH remaining en W12"
        cell.font = TAG_FONT; cell.fill = TAG_FILL; cell.alignment = LFT
        ws.row_dimensions[row].height = 18; row += 1

        for sub_name, acts in tag_data["subsistemas"].items():
            hh_sub = sum(float(a[4]) for a in acts)
            # Fila subsistema
            ws.merge_cells(f"A{row}:J{row}")
            cell = ws[f"A{row}"]
            cell.value = f"    {sub_name}  ({hh_sub:.1f} HH)"
            cell.font = SUB_FONT; cell.fill = SUB_FILL; cell.alignment = LFT
            ws.row_dimensions[row].height = 16; row += 1

            for act in acts:
                cod, nom, tgt, act_hh, rem, pct, estado, eend = act
                is_crit = estado == "NO INICIO" and rem >= 54
                flll = RED_FILL if is_crit else (WARN_FILL if pct < 50 and estado == "EN CURSO" else (OK_FILL if pct >= 75 else None))
                pct_fnt = RED_FONT if pct == 0 and rem >= 54 else (GRN_FONT if pct >= 75 else NRM_FONT)

                c(ws, row, 1, tag_name, flll=flll, aln=LFT)
                c(ws, row, 2, sub_name, flll=flll, aln=LFT)
                c(ws, row, 3, cod,      flll=flll)
                c(ws, row, 4, nom,      flll=flll, aln=LFT)
                c(ws, row, 5, tgt,      flll=flll)
                c(ws, row, 6, act_hh,   flll=flll)
                c(ws, row, 7, rem,      flll=flll)
                c(ws, row, 8, f"{pct}%", fnt=pct_fnt, flll=flll)
                c(ws, row, 9, estado,   fnt=pct_fnt, flll=flll)
                c(ws, row,10, eend,     flll=flll)
                ws.row_dimensions[row].height = 16
                row += 1

        row += 1  # espacio entre TAGs

wb = openpyxl.load_workbook(FILE)
add_w12(wb)
wb.save(FILE)
import sys; sys.stdout.reconfigure(encoding='utf-8')
print("OK: W12 - Plan de Trabajo agregado")
print("Hojas:", wb.sheetnames)
