"""
Agrega hojas de análisis al Excel EVM existente (sin tocar la hoja del usuario)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r'C:\Users\josej\Desktop\EVM_OT1844_W11.xlsx'

# ── Estilos ───────────────────────────────────────────────────────────────────
def fill(hex): return PatternFill("solid", fgColor=hex)
def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)
def brd():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HDR_FILL = fill("1F3864"); HDR_FONT = font(bold=True, color="FFFFFF", size=10)
INFO_FILL= fill("CFE2F3")
WARN_FILL= fill("FCE5CD")
OK_FILL  = fill("D9EAD3")
RED_FONT = font(bold=True, color="CC0000", size=10)
GRN_FONT = font(bold=True, color="006600", size=10)
NRM_FONT = font(size=10)
BLD_FONT = font(bold=True, size=10)

def cell(ws, row, col, val, fnt=None, flll=None, aln=CTR, border=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = fnt  or NRM_FONT
    c.alignment = aln
    if flll: c.fill = flll
    if border: c.border = brd()
    return c

def hdr_row(ws, row, cols, widths):
    for i,(h,w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CTR; c.border=brd()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28

def section_row(ws, row, ncols, text, bg="2E75B6"):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value=text; c.font=font(bold=True,size=10,color="FFFFFF")
    c.fill=fill(bg); c.alignment=LFT
    ws.row_dimensions[row].height=18

# ── Datos actividades en curso ────────────────────────────────────────────────
EN_CURSO = [
    ("A2830","Trazado y Corte de FRP",               54, 32.4, 21.6, 60.0, "17-Mar"),
    ("A3150","Trazado y Corte de FRP",               54, 32.4, 21.6, 60.0, "17-Mar"),
    ("A3280","Corte y Fab. Vigas y Perfil Z",       162,135.0, 27.0, 83.3, "16-Mar"),
    ("A3290","Corte y Fab. Consolas Exteriores",     36, 18.0, 18.0, 50.0, "16-Mar"),
    ("A3300","Corte y Fab. Consolas Centrales",      36, 18.0, 18.0, 50.0, "16-Mar"),
    ("A3420","Fab. Baranda Radial Removible",       189, 94.5, 94.5, 50.0, "19-Mar"),
    ("A3450","Trazado y Corte de FRP",               54, 37.8, 16.2, 70.0, "20-Mar"),
    ("A3470","Fab. Cuerpo Central (Spool/Cabezal)",  36, 28.8,  7.2, 80.0, "16-Mar"),
    ("A3580","Corte y Fab. Vigas y Perfil Z",       162, 81.0, 81.0, 50.0, "18-Mar"),
    ("A3590","Corte y Fab. Consolas Exteriores",     36, 10.8, 25.2, 30.0, "17-Mar"),
    ("A3600","Corte y Fab. Consolas Centrales",      36, 10.8, 25.2, 30.0, "18-Mar"),
    ("A3720","Fab. Baranda Radial Removible",       189, 94.5, 94.5, 50.0, "19-Mar"),
    ("A3750","Trazado y Corte de FRP",               54, 37.8, 16.2, 70.0, "20-Mar"),
    ("A3770","Fab. Cuerpo Central (Spool/Cabezal)",  36, 28.8,  7.2, 80.0, "16-Mar"),
    ("A4050","Trazado y Corte de FRP",               54, 37.8, 16.2, 70.0, "27-Mar"),
    ("A4070","Fab. Cuerpo Central (Spool/Cabezal)",  36, 28.8,  7.2, 80.0, "23-Mar"),
]

# ── Hoja 2: Estado de Actividades ────────────────────────────────────────────
def add_estado(wb):
    ws = wb.create_sheet("Estado Actividades")
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:G1")
    c=ws["A1"]; c.value="ESTADO DE ACTIVIDADES — OT-1844 | Corte W11 | 15-Mar-2026"
    c.font=font(bold=True,size=12,color="1F3864"); c.alignment=CTR; c.fill=INFO_FILL
    ws.row_dimensions[1].height=22

    ws.merge_cells("A2:G2")
    c=ws["A2"]
    c.value=("  Completadas: 68 act | 3,942 HH (48.67% BAC)   "
             "|   En Curso: 16 act | 727.2 HH target   "
             "|   No Iniciadas: 68 act | 3,430.8 HH remaining")
    c.font=BLD_FONT; c.alignment=LFT; c.fill=INFO_FILL
    ws.row_dimensions[2].height=16

    hdrs=["Código","Actividad","HH Target","HH Act.","HH Rem.","Avance %","Early End"]
    wids=[10, 48, 11, 11, 11, 11, 12]
    hdr_row(ws, 3, hdrs, wids)

    # EN CURSO
    row=4
    section_row(ws, row, 7, "  EN CURSO — 16 actividades | 727.2 HH ejecutadas parcialmente")
    row+=1

    for cod,nom,tgt,act_hh,rem,pct,eend in EN_CURSO:
        flll = WARN_FILL if pct < 50 else (OK_FILL if pct >= 75 else None)
        pct_fnt = RED_FONT if pct < 50 else (GRN_FONT if pct >= 75 else NRM_FONT)
        cell(ws,row,1,cod,  flll=flll)
        cell(ws,row,2,nom,  flll=flll, aln=LFT)
        cell(ws,row,3,tgt,  flll=flll)
        cell(ws,row,4,act_hh,flll=flll)
        cell(ws,row,5,rem,  flll=flll)
        cell(ws,row,6,f"{pct}%", fnt=pct_fnt, flll=flll)
        cell(ws,row,7,eend, flll=flll)
        ws.row_dimensions[row].height=16
        row+=1

    row+=1
    section_row(ws, row, 7,
        "  COMPLETADAS — 68 actividades | 3,942 HH (100% ejecutadas al corte)", bg="375623")
    row+=2

    section_row(ws, row, 7,
        "  NO INICIADAS — 68 actividades | 3,430.8 HH remaining | Arrancan desde 16-Mar-2026", bg="7F7F7F")

# ── Hoja 3: Análisis de Control ──────────────────────────────────────────────
def add_analisis(wb):
    ws = wb.create_sheet("Análisis de Control")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 65

    ws.merge_cells("A1:B1")
    c=ws["A1"]; c.value="ANÁLISIS DE CONTROL — OT-1844 | W11 | Corte 15-Mar-2026"
    c.font=font(bold=True,size=12,color="1F3864"); c.alignment=CTR; c.fill=INFO_FILL
    ws.row_dimensions[1].height=22

    rows_data = [
        ("INDICADORES EVM", None),
        ("BAC",                      "8,100 HH"),
        ("EV al corte (15-Mar)",     "4,669.2 HH  →  57.64% del BAC"),
        ("PV planificado W11",       "5,283.0 HH  →  65.22% del BAC"),
        ("SV (Variación Cronograma)","−613.8 HH  (atraso respecto al plan)"),
        ("SPI",                      "0.884  —  por cada HH planificada se ejecuta 0.88 HH"),
        ("Remaining (ETC)",          "3,430.8 HH"),
        ("EAC (Forecast al cierre)", "8,100.0 HH  —  cierra en W15 si se mantiene ritmo actual"),
        (None, None),
        ("TENDENCIA SPI", None),
        ("Evolución semanal",        "W08: 0.898  →  W09: 0.834  →  W10: 0.810  →  W11: 0.884"),
        ("Interpretación",           "Leve recuperación en W11. SPI aún bajo 1.0 — ritmo insuficiente para recuperar atraso sin aceleración."),
        (None, None),
        ("ACTIVIDADES CRÍTICAS", None),
        ("Bajo avance al corte",     "A3590 / A3600 — Consolas Exteriores y Centrales al 30% (Early End 16-17 Mar)"),
        ("Baranda radial",           "A3420 / A3720 — Fab. Baranda Radial al 50% (189 HH target c/u, Early End 19-Mar)"),
        ("Vigas en atraso",          "A3580 — Corte y Fab. Vigas y Perfil Z al 50% (162 HH, Early End 18-Mar)"),
        ("FRP en curso",             "A2830 / A3150 / A3450 / A3750 / A4050 — Trazado y Corte FRP entre 60-70%"),
        (None, None),
        ("RIESGO PRINCIPAL", None),
        ("Actividades no iniciadas", "68 actividades con 3,430.8 HH remaining arrancan desde 16-Mar — ventana de ejecución muy estrecha"),
        ("Riesgo de cierre",         "Si no se inician todas las actividades pendientes el 16-Mar, el forecast W15 se compromete"),
        ("Material SS316L",          "Actividades de armado/soldadura SS316L no iniciadas (A3390, A3530, A3690, A3830) con Early Start 16-Mar"),
        (None, None),
        ("FORECAST", None),
        ("W12 (16–22 Mar)",          "6,541.2 HH acum  →  80.76%  |  RE semana: 1,872.0 HH"),
        ("W13 (23–29 Mar)",          "7,539.3 HH acum  →  93.08%  |  RE semana: 998.1 HH"),
        ("W14 (30 Mar–05 Abr)",      "8,064.0 HH acum  →  99.56%  |  RE semana: 524.7 HH"),
        ("W15 (06–12 Abr)",          "8,100.0 HH acum  →  100.00%  ✅  Cierre proyectado en plazo"),
        (None, None),
        ("CONCLUSIÓN", None),
        ("Estado general",
         "ATENCIÓN — SPI 0.884, atraso de 7.6 puntos al corte W11. "
         "La recuperación es posible si se inician las 68 actividades pendientes sin demora "
         "a partir del 16-Mar y se mantiene el ritmo de Remaining Early proyectado. "
         "Seguimiento crítico requerido en W12."),
    ]

    for r,(k,v) in enumerate(rows_data, 2):
        ws.row_dimensions[r].height = 18
        if k is None:
            continue
        if v is None:
            # Sección
            ws.merge_cells(f"A{r}:B{r}")
            c=ws[f"A{r}"]; c.value=k
            c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=LFT
        else:
            c1=ws.cell(row=r,column=1,value=k)
            c1.font=BLD_FONT; c1.alignment=LFT; c1.border=brd()
            c2=ws.cell(row=r,column=2,value=v)
            c2.font=NRM_FONT; c2.alignment=LFT; c2.border=brd()
            if "ATENCIÓN" in v:
                c2.fill=WARN_FILL; c2.font=RED_FONT
            elif "✅" in v:
                c2.fill=OK_FILL
            elif "30%" in v or "bajo" in k.lower() or "riesgo" in k.lower():
                c2.fill=WARN_FILL

# ── Main ──────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILE)

# Eliminar hojas previas si existen (para regenerar limpias)
for sh in ["Estado Actividades", "Análisis de Control"]:
    if sh in wb.sheetnames:
        del wb[sh]

add_estado(wb)
add_analisis(wb)
wb.save(FILE)
print(f"✅ Guardado: {FILE}")
print(f"Hojas: {wb.sheetnames}")
